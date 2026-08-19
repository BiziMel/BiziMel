import os
import re
import base64
import io
import json
import sqlite3
import tempfile
from datetime import date, datetime, time, timedelta
from pathlib import Path


def assert_ok(condition, message):
    if not condition:
        raise AssertionError(message)


def csrf_from_session(client):
    with client.session_transaction() as sess:
        return sess.get("_csrf_token", "")


def input_value(html, field_id):
    match = re.search(rf'id="{re.escape(field_id)}"[^>]*\bvalue="([^"]*)"', html)
    return match.group(1) if match else ""


def seed_validation_data(db_path):
    connection = sqlite3.connect(db_path)
    account_id = connection.execute(
        """
        INSERT INTO accounts (
            account_name, pg_bible_order, account_tier, industry, business_unit,
            country, city, website, pipeline_target, notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "Smoke Test Account",
            1,
            "1",
            "Public Sector",
            "BMC",
            "United Kingdom",
            "London",
            "https://example.com",
            500000,
            "Smoke test notes",
        ),
    ).lastrowid
    smoke_logo = (
        "data:image/png;base64,"
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
    )
    connection.execute(
        "UPDATE accounts SET customer_logo = ? WHERE id = ?",
        (smoke_logo, account_id),
    )
    connection.execute(
        "UPDATE accounts SET pipeline_target = ?, current_pipeline = ?, account_tier = ? WHERE id = ?",
        ("£1,250,000", "$250,000", "Tier 1", account_id),
    )
    sales_play_id = connection.execute(
        """
        INSERT INTO sales_plays (sales_play_title, sales_play_description, sales_play_products)
        VALUES (?, ?, ?)
        """,
        ("Smoke Test Play", "Configured smoke Sales Play", "Smoke Product"),
    ).lastrowid
    connection.execute(
        """
        INSERT INTO account_sales_plays (account_id, sales_play_id)
        VALUES (?, ?)
        """,
        (account_id, sales_play_id),
    )
    contact_id = connection.execute(
        """
        INSERT INTO contacts (account_id, category, name, job_title, email, phone, location)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (account_id, "Executive", "Smoke Test Contact", "CIO", "contact@example.com", "12345", "London"),
    ).lastrowid
    second_contact_id = connection.execute(
        """
        INSERT INTO contacts (account_id, category, name, job_title, email, phone, location)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (account_id, "Technical", "Smoke Second Contact", "CTO", "second@example.com", "67890", "London"),
    ).lastrowid
    partner_id = connection.execute(
        """
        INSERT INTO partners (partner_name, partner_type, country, city, relationship_owner)
        VALUES (?, ?, ?, ?, ?)
        """,
        ("Smoke Test Partner", "SI", "United Kingdom", "London", "Smoke Tester"),
    ).lastrowid
    partner_contact_id = connection.execute(
        """
        INSERT INTO partner_contacts (partner_id, account_id, name, job_title, partner_contact_role, email)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (partner_id, account_id, "Smoke Partner Contact", "CTO", "Presales", "partner@example.com"),
    ).lastrowid
    connection.execute(
        """
        INSERT INTO account_partners (account_id, partner_id, partner_name, partner_role, involvement_status)
        VALUES (?, ?, ?, ?, ?)
        """,
        (account_id, partner_id, "Smoke Test Partner", "Account Manager", "Active"),
    )
    outreach_id = connection.execute(
        """
        INSERT INTO outreach (
            fy, quarter, account_id, contact_id, campaign, sales_play, campaign_start_date,
            campaign_end_date, campaign_tasks_per_week, campaign_total_tasks, activity_date,
            activity_time, activity_type, subject, notes, outcome, next_action,
            next_action_date, next_action_time, task_status, assigned_to
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "FY27",
            "Q1",
            account_id,
            contact_id,
            "VITO",
            "Smoke Test Play",
            "2026-05-01",
            "2026-05-29",
            3,
            12,
            "2026-05-01",
            "09:00",
            "Call",
            "Smoke test outreach",
            "Sales play: Smoke Test Play. Contact: Smoke Test Contact",
            "Meeting Booked",
            "Follow up",
            "2026-05-05",
            "10:00",
            "Not Started",
            "Smoke Tester",
        ),
    ).lastrowid
    connection.execute(
        """
        UPDATE outreach
        SET scheduled_meeting_date = ?,
            scheduled_meeting_time = ?
        WHERE id = ?
        """,
        ("2026-05-06", "10:30", outreach_id),
    )
    connection.execute(
        """
        INSERT INTO outreach (
            fy, quarter, account_id, contact_id, campaign, sales_play, campaign_start_date,
            campaign_end_date, campaign_tasks_per_week, campaign_total_tasks, activity_date,
            activity_time, activity_type, subject, notes, outcome, next_action,
            next_action_date, next_action_time, scheduled_meeting_date, scheduled_meeting_time,
            task_status, assigned_to
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "FY27",
            "Q1",
            account_id,
            contact_id,
            "NBM Smoke Campaign",
            "Smoke Test Play",
            "2026-05-01",
            "2026-05-29",
            1,
            1,
            "2026-05-04",
            "09:30",
            "Meeting",
            "NBM booked smoke",
            "NBM booked smoke coverage",
            "NBM Booked",
            "Attend NBM",
            "2026-05-08",
            "10:30",
            "2026-05-08",
            "10:30",
            "Not Started",
            "Smoke Tester",
        ),
    )
    connection.execute(
        """
        INSERT INTO outreach (
            fy, quarter, account_id, contact_id, campaign, sales_play, campaign_start_date,
            campaign_end_date, campaign_tasks_per_week, campaign_total_tasks, activity_date,
            activity_time, activity_type, subject, notes, outcome, next_action,
            next_action_date, next_action_time, task_status, assigned_to
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "FY27",
            "Q1",
            account_id,
            contact_id,
            "Deleted Smoke Campaign",
            "Deleted PG Progress Play",
            "2026-05-01",
            "2026-05-29",
            1,
            1,
            "2026-07-29",
            "11:00",
            "Email",
            "Deleted PG Progress Marker",
            "This deleted activity must not appear in PG Progress",
            "Meeting Booked",
            "Deleted PG Progress Marker",
            "2026-08-05",
            "11:30",
            "Deleted",
            "Smoke Tester",
        ),
    )
    connection.execute(
        """
        INSERT INTO outreach (
            fy, quarter, account_id, contact_id, campaign, sales_play, campaign_start_date,
            campaign_end_date, campaign_tasks_per_week, campaign_total_tasks, activity_date,
            activity_time, activity_type, subject, notes, outcome, next_action,
            next_action_date, next_action_time, task_status, assigned_to
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "FY27",
            "Q1",
            account_id,
            second_contact_id,
            "Older Smoke Campaign",
            "Smoke Test Play",
            "2026-05-01",
            "2026-05-29",
            1,
            1,
            "2026-05-01",
            "12:00",
            "Email",
            "Older PG Progress Marker",
            "Older activity should still keep this contact visible in PG Progress",
            "No Response",
            "Review older outreach route",
            "2026-05-02",
            "12:30",
            "Completed",
            "Smoke Tester",
        ),
    )
    connection.commit()
    connection.close()
    return account_id, contact_id, second_contact_id, partner_id, partner_contact_id, outreach_id


def main():
    with tempfile.TemporaryDirectory() as tmp:
        os.environ["PIPEFLOW_DATA_DIR"] = tmp
        os.environ.pop("DATABASE_URL", None)
        os.environ["PIPEFLOW_SECRET_KEY"] = "pipeflow-smoke-test-key"

        import app as pipeflow_app
        import db_compat

        assert_ok(
            {"outreach_recipients", "audit_entries"}.issubset(db_compat.USER_TABLES),
            "tenant workspace tables missing from Postgres compatibility layer",
        )
        retry_attempts = {"count": 0}
        retry_rollbacks = {"count": 0}

        def aborted_transaction_then_success():
            retry_attempts["count"] += 1
            if retry_attempts["count"] == 1:
                raise RuntimeError("current transaction is aborted, commands ignored until end of transaction block")
            return "recovered"

        retry_result = db_compat.execute_with_retry(
            aborted_transaction_then_success,
            rollback=lambda: retry_rollbacks.__setitem__("count", retry_rollbacks["count"] + 1),
        )
        assert_ok(
            retry_result == "recovered" and retry_rollbacks["count"] == 1,
            "Postgres aborted transaction recovery did not rollback and retry",
        )
        assert_ok(
            len(pipeflow_app.diagnostic_error_code("OUTREACH-ADD")) <= 10,
            "diagnostic error code is too long",
        )
        today = pipeflow_app.current_app_datetime().date()
        calc = pipeflow_app.calculate_automated_pg_rag_status
        meeting_future = today + timedelta(days=7)
        stale_amber = today - timedelta(days=15)
        stale_red = today - timedelta(days=31)
        assert_ok(
            calc([{"outcome": "Positive Response", "scheduled_meeting_date": "", "task_status": "Completed", "completed_at": today.isoformat()}], today=today)["automatedRagStatus"] == "red",
            "PG RAG should stay red for a positive response without a scheduled booked meeting",
        )
        assert_ok(
            calc([{"outcome": "NBM Booked", "scheduled_meeting_date": "", "task_status": "Not Started", "next_action_date": meeting_future.isoformat()}], today=today)["automatedRagStatus"] == "red",
            "PG RAG should stay red when a meeting outcome has no scheduled meeting date",
        )
        assert_ok(
            calc([{"outcome": "No Response", "scheduled_meeting_date": meeting_future.isoformat(), "task_status": "Not Started", "next_action_date": meeting_future.isoformat()}], today=today)["automatedRagStatus"] == "red",
            "PG RAG should stay red when a scheduled date exists without a booked meeting outcome",
        )
        assert_ok(
            calc([{"outcome": "NBM Booked", "scheduled_meeting_date": meeting_future.isoformat(), "task_status": "Not Started", "next_action_date": meeting_future.isoformat()}], today=today)["automatedRagStatus"] == "green",
            "PG RAG should be green for a booked meeting with a valid future scheduled meeting date",
        )
        assert_ok(
            calc([{"outcome": "NBM Booked", "scheduled_meeting_date": stale_amber.isoformat(), "task_status": "Completed", "completed_at": stale_amber.isoformat()}], today=today)["automatedRagStatus"] == "amber",
            "PG RAG should degrade to amber after 14 days without scheduled or closed activity",
        )
        assert_ok(
            calc([{"outcome": "NBM Booked", "scheduled_meeting_date": stale_red.isoformat(), "task_status": "Completed", "completed_at": stale_red.isoformat()}], today=today)["automatedRagStatus"] == "red",
            "PG RAG should degrade to red after 30 days without scheduled or closed activity",
        )
        assert_ok(
            pipeflow_app.effective_pg_rag_payload({"automatedRagStatus": "red", "reason": "auto"}, "green")["effectiveRagStatus"] == "green",
            "PG RAG manual override should take precedence over automatic status",
        )

        client = pipeflow_app.app.test_client()
        for path in ("/login", "/register", "/forgot-password"):
            response = client.get(path)
            assert_ok(response.status_code == 200, f"{path} returned {response.status_code}")

        future_meeting_date = (today + timedelta(days=14)).isoformat()
        response = client.post(
            "/register",
            data={
                "csrf_token": csrf_from_session(client),
                "full_name": "Smoke Test Admin",
                "email": "smoke-test@example.com",
                "password": "Password123!",
                "reset_phrase": "smoke test secret phrase",
            },
            follow_redirects=True,
        )
        assert_ok(response.status_code == 200, f"register returned {response.status_code}")
        version_response = client.get("/health/version")
        assert_ok(
            version_response.status_code == 200
            and "pipeflow_version=2.8.1" in version_response.get_data(as_text=True),
            "health/version did not report Release 2.8.1",
        )

        response = client.post(
            "/logout",
            data={"csrf_token": csrf_from_session(client)},
            follow_redirects=True,
        )
        assert_ok(response.status_code == 200 and "Sign In" in response.get_data(as_text=True), "logout failed")
        response = client.post(
            "/login",
            data={
                "csrf_token": "stale-token-after-deploy",
                "email": "smoke-test@example.com",
                "password": "Password123!",
            },
            follow_redirects=True,
        )
        assert_ok(response.status_code == 200 and "Dashboard" in response.get_data(as_text=True), "stale-token login recovery failed")

        db_path = Path(tmp) / "users" / "1" / "pipeflow.db"
        account_id, contact_id, second_contact_id, partner_id, partner_contact_id, outreach_id = seed_validation_data(db_path)
        today = date.today()
        yesterday = (today - timedelta(days=1)).isoformat()
        tomorrow = (today + timedelta(days=1)).isoformat()

        failed_job_date = today - timedelta(days=5)
        failed_job_now = datetime.combine(failed_job_date, time(23, 0))
        job_token = pipeflow_app.claim_scheduled_job(
            "nightly_outreach_schedule",
            failed_job_date,
            failed_job_now,
        )
        assert_ok(job_token, "nightly scheduler run could not be claimed")
        assert_ok(
            not pipeflow_app.claim_scheduled_job("nightly_outreach_schedule", failed_job_date, failed_job_now),
            "nightly scheduler allowed the same daily run to be claimed twice",
        )
        pipeflow_app.finish_scheduled_job(
            "nightly_outreach_schedule",
            failed_job_date,
            job_token,
            "failed",
            "Smoke scheduler failure",
            failed_job_now,
        )
        failed_alert_html = client.get("/").get_data(as_text=True)
        assert_ok(
            "Nightly schedule review needs attention" in failed_alert_html
            and "Smoke scheduler failure" in failed_alert_html,
            "administrators were not shown the nightly service failure dialog",
        )
        auth_connection = pipeflow_app.get_auth_connection()
        auth_connection.execute(
            "UPDATE scheduled_job_runs SET status = 'completed', detail = 'Smoke recovery complete' WHERE job_key = ?",
            (f"nightly_outreach_schedule:{failed_job_date.isoformat()}",),
        )
        auth_connection.commit()
        auth_connection.close()

        pages = {
            "/": "Dashboard",
            "/accounts": "Accounts",
            f"/accounts/{account_id}": "Smoke Test Account",
            "/contacts": "Contacts",
            f"/contacts/{contact_id}": "Smoke Test Contact",
            "/partners": "Partners",
            f"/partners/{partner_id}": "Smoke Test Partner",
            "/outreach": "Outreach",
            f"/outreach/{outreach_id}": "Smoke test outreach",
            "/outreach/add": "Add Outreach",
            "/outreach/campaign-builder": "Campaign Builder",
            "/reports": "Reports",
            "/reports/accounts": "Account Reports",
            "/reports/contacts": "Contact Reports",
            "/reports/outreach": "Outreach Reports",
            "/reports/partners": "Partner Reports",
            "/reports/sales-plays": "Sales Play Reports",
            "/reports/pg-progress": "PG Progress",
            "/reports/tasks": "Outreach Reports",
            "/search?q=Smoke": "Global Search",
            "/profile": "Profile",
            "/admin/permissions": "Admin",
        }
        for path, marker in pages.items():
            response = client.get(path)
            html = response.get_data(as_text=True)
            assert_ok(response.status_code == 200 and marker in html, f"{path} failed")

        add_outreach_html = client.get("/outreach/add").get_data(as_text=True)
        campaign_builder_html = client.get("/outreach/campaign-builder").get_data(as_text=True)
        assert_ok("SMS/WhatsApp" in add_outreach_html, "Add Outreach activity type list missing SMS/WhatsApp")
        assert_ok("SMS/WhatsApp" in campaign_builder_html, "Campaign Builder activity type list missing SMS/WhatsApp")

        availability_date = today + timedelta(days=120)
        while availability_date.weekday() >= 5:
            availability_date += timedelta(days=1)
        parsed_part_day_block = pipeflow_app.parse_non_working_blocks([{
            "start_date": availability_date.isoformat(),
            "start_time": "11:45",
            "end_date": availability_date.isoformat(),
            "end_time": "13:00",
        }])[0]
        assert_ok(
            pipeflow_app.datetime_within_blocks(
                datetime.combine(availability_date, datetime.strptime("12:15", "%H:%M").time()),
                [parsed_part_day_block],
            ),
            "part-day non-working interval did not block an unavailable time",
        )
        assert_ok(
            not pipeflow_app.datetime_within_blocks(
                datetime.combine(availability_date, datetime.strptime("13:00", "%H:%M").time()),
                [parsed_part_day_block],
            ),
            "part-day non-working interval incorrectly blocked its return time",
        )
        campaign_slot_date, campaign_slot_time = pipeflow_app.next_campaign_working_slot(
            availability_date,
            availability_date,
            "12:00",
            profile={"work_day_start": "09:00", "work_day_end": "17:00"},
            reserved_slots=set(),
            non_working_blocks=[parsed_part_day_block],
        )
        assert_ok(
            campaign_slot_date == availability_date and campaign_slot_time == "13:00",
            "Campaign Builder scheduling did not move beyond a part-day absence",
        )

        connection = sqlite3.connect(db_path)
        availability_task_ids = []
        for index, (scheduled_time, selected_contact_id) in enumerate(
            (("12:00", contact_id), ("12:15", second_contact_id), ("13:00", contact_id)),
            start=1,
        ):
            cursor = connection.execute(
                """
                INSERT INTO outreach (
                    fy, quarter, account_id, contact_id, campaign, sales_play,
                    activity_date, activity_time, activity_type, subject, outcome,
                    next_action, next_action_date, next_action_time, task_status, assigned_to
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "27", "Q1", account_id, selected_contact_id, "Availability smoke", "Smoke Test Play",
                    availability_date.isoformat(), scheduled_time, "Email",
                    f"Availability ordered task {index}", "No Response", "Continue engagement",
                    availability_date.isoformat(), scheduled_time, "Not Started", "Smoke Test Admin",
                ),
            )
            availability_task_ids.append(cursor.lastrowid)
        connection.commit()
        connection.close()
        response = client.post(
            "/profile/non-working/add",
            data={
                "csrf_token": csrf_from_session(client),
                "start_date": availability_date.isoformat(),
                "start_time": "11:45",
                "end_date": availability_date.isoformat(),
                "end_time": "13:00",
                "reason": "Part-day smoke absence",
            },
            follow_redirects=True,
        )
        availability_html = response.get_data(as_text=True)
        assert_ok(
            response.status_code == 200
            and "3 open Outreach task(s) were rescheduled from the first conflict onward" in availability_html
            and "From Time" in availability_html
            and "To Time" in availability_html,
            "part-day non-working block was not saved or reported clearly",
        )
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        moved_rows = connection.execute(
            "SELECT id, activity_date, activity_time, next_action_date, next_action_time FROM outreach WHERE id IN (?, ?, ?)",
            tuple(availability_task_ids),
        ).fetchall()
        moved_rows_by_id = {row["id"]: row for row in moved_rows}
        first_moved = datetime.fromisoformat(
            f"{moved_rows_by_id[availability_task_ids[0]]['next_action_date']}T{moved_rows_by_id[availability_task_ids[0]]['next_action_time']}"
        )
        second_moved = datetime.fromisoformat(
            f"{moved_rows_by_id[availability_task_ids[1]]['next_action_date']}T{moved_rows_by_id[availability_task_ids[1]]['next_action_time']}"
        )
        third_moved = datetime.fromisoformat(
            f"{moved_rows_by_id[availability_task_ids[2]]['next_action_date']}T{moved_rows_by_id[availability_task_ids[2]]['next_action_time']}"
        )
        assert_ok(first_moved < second_moved < third_moved, "availability rescheduling did not preserve the original task order")
        assert_ok(
            all(not pipeflow_app.datetime_within_blocks(value, [parsed_part_day_block]) for value in (first_moved, second_moved, third_moved)),
            "availability rescheduling left a task inside the new part-day absence",
        )
        expired_date = (today - timedelta(days=3)).isoformat()
        expired_cursor = connection.execute(
            "INSERT INTO non_working_blocks (start_date, start_time, end_date, end_time, reason) VALUES (?, ?, ?, ?, ?)",
            (expired_date, "09:00", expired_date, "10:00", "Expired smoke block"),
        )
        expired_block_id = expired_cursor.lastrowid
        connection.commit()
        connection.close()
        client.get("/profile")
        connection = sqlite3.connect(db_path)
        expired_count = connection.execute(
            "SELECT COUNT(*) FROM non_working_blocks WHERE id = ?",
            (expired_block_id,),
        ).fetchone()[0]
        assert_ok(expired_count == 0, "expired profile non-working block was not removed automatically")
        connection.execute(
            "DELETE FROM outreach WHERE id IN (?, ?, ?)",
            tuple(availability_task_ids),
        )
        connection.execute("DELETE FROM non_working_blocks WHERE reason = ?", ("Part-day smoke absence",))
        connection.commit()
        connection.close()

        nightly_block_date = today + timedelta(days=220)
        while nightly_block_date.weekday() != 4:
            nightly_block_date += timedelta(days=1)
        nightly_now = datetime.combine(
            nightly_block_date - timedelta(days=1),
            datetime.strptime("23:00", "%H:%M").time(),
        )
        nightly_original_slots = [
            datetime.combine(nightly_block_date, datetime.strptime("10:00", "%H:%M").time()),
            datetime.combine(nightly_block_date, datetime.strptime("10:15", "%H:%M").time()),
            datetime.combine(nightly_block_date + timedelta(days=1), datetime.strptime("10:30", "%H:%M").time()),
        ]
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        connection.execute(
            "INSERT INTO non_working_blocks (start_date, start_time, end_date, end_time, reason) VALUES (?, ?, ?, ?, ?)",
            (nightly_block_date.isoformat(), "10:00", nightly_block_date.isoformat(), "12:00", "Nightly review smoke block"),
        )
        nightly_task_ids = []
        for index, original_slot in enumerate(nightly_original_slots):
            cursor = connection.execute(
                """
                INSERT INTO outreach (
                    fy, quarter, account_id, contact_id, campaign, sales_play,
                    activity_date, activity_time, activity_type, subject, outcome,
                    next_action, next_action_date, next_action_time, task_status, assigned_to
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "27", "Q1", account_id, second_contact_id, "Nightly scheduler smoke", "Smoke Test Play",
                    original_slot.date().isoformat(), original_slot.strftime("%H:%M"), "Email",
                    f"Nightly ordered task {index + 1}", "", "Continue engagement",
                    original_slot.date().isoformat(), original_slot.strftime("%H:%M"), "Not Started", "Smoke Test Admin",
                ),
            )
            nightly_task_ids.append(cursor.lastrowid)
        closed_slot = nightly_original_slots[0]
        closed_cursor = connection.execute(
            """
            INSERT INTO outreach (
                fy, quarter, account_id, contact_id, campaign, sales_play,
                activity_date, activity_time, activity_type, subject, outcome,
                next_action, next_action_date, next_action_time, task_status, assigned_to
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "27", "Q1", account_id, contact_id, "Nightly scheduler smoke", "Smoke Test Play",
                closed_slot.date().isoformat(), closed_slot.strftime("%H:%M"), "Email",
                "Nightly closed history", "Positive Response", "", closed_slot.date().isoformat(),
                closed_slot.strftime("%H:%M"), "Closed", "Smoke Test Admin",
            ),
        )
        closed_task_id = closed_cursor.lastrowid
        nightly_test_ids = {*nightly_task_ids, closed_task_id}
        other_task_statuses = connection.execute(
            "SELECT id, task_status FROM outreach WHERE id NOT IN (?, ?, ?, ?)",
            tuple(nightly_test_ids),
        ).fetchall()
        connection.execute(
            "UPDATE outreach SET task_status = 'Closed' WHERE id NOT IN (?, ?, ?, ?)",
            tuple(nightly_test_ids),
        )
        connection.commit()
        moved_count = pipeflow_app.nightly_reflow_outreach_schedule(connection, nightly_now)
        connection.commit()
        assert_ok(moved_count == 3, "nightly review did not move every invalid open scheduled task")
        nightly_rows = connection.execute(
            "SELECT id, next_action_date, next_action_time FROM outreach WHERE id IN (?, ?, ?) ORDER BY next_action_date, next_action_time, id",
            tuple(nightly_task_ids),
        ).fetchall()
        nightly_moved_slots = [
            datetime.fromisoformat(f"{row['next_action_date']}T{row['next_action_time']}")
            for row in nightly_rows
        ]
        assert_ok(
            nightly_moved_slots == sorted(nightly_moved_slots) and len(set(nightly_moved_slots)) == 3,
            "nightly review did not preserve order in distinct schedule slots",
        )
        assert_ok(
            all(moved >= original for moved, original in zip(nightly_moved_slots, nightly_original_slots)),
            "nightly review moved an Outreach task backwards",
        )
        assert_ok(
            all(slot.weekday() < 5 for slot in nightly_moved_slots)
            and all(not (slot.date() == nightly_block_date and time(10, 0) <= slot.time() < time(12, 0)) for slot in nightly_moved_slots),
            "nightly review left Outreach inside a weekend or blocked time",
        )
        closed_row = connection.execute(
            "SELECT next_action_date, next_action_time FROM outreach WHERE id = ?",
            (closed_task_id,),
        ).fetchone()
        assert_ok(
            closed_row["next_action_date"] == closed_slot.date().isoformat() and closed_row["next_action_time"] == closed_slot.strftime("%H:%M"),
            "nightly review changed closed Outreach history",
        )
        connection.commit()
        second_pass_count = pipeflow_app.nightly_reflow_outreach_schedule(connection, nightly_now)
        assert_ok(second_pass_count == 0, "nightly review was not idempotent for an already valid schedule")
        connection.execute(
            "DELETE FROM outreach WHERE id IN (?, ?, ?, ?)",
            (*nightly_task_ids, closed_task_id),
        )
        for task_row in other_task_statuses:
            connection.execute(
                "UPDATE outreach SET task_status = ? WHERE id = ?",
                (task_row["task_status"], task_row["id"]),
            )
        connection.execute("DELETE FROM non_working_blocks WHERE reason = ?", ("Nightly review smoke block",))
        connection.commit()
        connection.close()

        outreach_report_html = client.get("/reports/outreach").get_data(as_text=True)
        assert_ok("Older PG Progress Marker" in outreach_report_html, "Outreach Reports did not show all records by default")
        assert_ok("name=\"company_id\"" in outreach_report_html, "Outreach Reports company/account filter missing")
        assert_ok("name=\"contact_id\"" in outreach_report_html, "Outreach Reports contact filter missing")
        assert_ok("name=\"last_updated_start\"" in outreach_report_html, "Outreach Reports last updated filter missing")
        assert_ok("name=\"task_status\"" in outreach_report_html, "Outreach Reports status filter missing")
        assert_ok("name=\"due_start_date\"" in outreach_report_html, "Outreach Reports due date filter missing")
        assert_ok("Last Updated" in outreach_report_html, "Outreach Reports table last updated column missing")
        filtered_export = client.get(
            f"/reports/outreach/export?company_id={account_id}&contact_id={second_contact_id}&activity_type=Email&due_start_date=2026-05-02&due_end_date=2026-05-02"
        ).get_data(as_text=True)
        assert_ok("Last Updated" in filtered_export, "Outreach Reports export missing Last Updated header")
        assert_ok("Review older outreach route" in filtered_export, "Outreach Reports export did not include matching filtered outreach")
        assert_ok("Follow up" not in filtered_export, "Outreach Reports export ignored selected filters")

        contacts_html = client.get("/contacts").get_data(as_text=True)
        assert_ok("Last Outreach" in contacts_html, "Contacts table last outreach column missing")
        assert_ok("08-05-2026 10:30" in contacts_html, "Contacts table did not show the latest active outreach date")
        assert_ok("05-08-2026 11:30" not in contacts_html, "Contacts table used a deleted outreach as the last active outreach date")

        accounts_html = client.get("/accounts").get_data(as_text=True)
        logo_path = f"/accounts/{account_id}/logo"
        assert_ok(logo_path in accounts_html, "Accounts table logo image source missing")
        logo_response = client.get(logo_path)
        assert_ok(logo_response.status_code == 200, "Account logo image route failed")
        assert_ok(logo_response.content_type.startswith("image/png"), "Account logo route returned the wrong content type")
        assert_ok(logo_response.get_data(), "Account logo route returned no image data")

        account_html = client.get(f"/accounts/{account_id}").get_data(as_text=True)
        add_contact_path = f"/contacts/add?account_id={account_id}"
        assert_ok(add_contact_path in account_html, "Account page Add Contact link is not account-specific")
        add_contact_html = client.get(add_contact_path).get_data(as_text=True)
        assert_ok(f'value="{account_id}"' in add_contact_html and "selected" in add_contact_html, "Add Contact account prefill missing")

        response = client.post(
            f"/accounts/{account_id}/org-chart/create",
            data={
                "csrf_token": csrf_from_session(client),
                "chart_name": "Smoke Manual Connector Chart",
                "notes": "Manual connector smoke coverage",
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "Org chart create failed")
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        chart_id = connection.execute(
            "SELECT id FROM account_org_charts WHERE account_id = ? ORDER BY id DESC LIMIT 1",
            (account_id,),
        ).fetchone()["id"]
        connection.close()
        org_chart_html = client.get(f"/accounts/{account_id}/org-chart?chart_id={chart_id}").get_data(as_text=True)
        assert_ok("Connectors" in org_chart_html and "org-connector-dot" in org_chart_html, "Manual org chart connector controls missing")
        assert_ok("click the first tile" in org_chart_html, "Org chart click-to-connect guidance missing")
        layout_actions = [
            {
                "type": "person",
                "local_node_id": "smoke-contact-1",
                "person_ref": f"contact:{contact_id}",
                "x_position": 32,
                "y_position": 32,
                "sort_order": 1,
            },
            {
                "type": "person",
                "local_node_id": "smoke-contact-2",
                "person_ref": f"contact:{second_contact_id}",
                "x_position": 288,
                "y_position": 32,
                "sort_order": 2,
            },
            {
                "type": "connector",
                "source_local_id": "smoke-contact-1",
                "target_local_id": "smoke-contact-2",
                "source_side": "right",
                "target_side": "left",
                "orientation": "horizontal",
            },
        ]
        response = client.post(
            f"/accounts/{account_id}/org-chart/{chart_id}/layout/save",
            data={
                "csrf_token": csrf_from_session(client),
                "layout_actions": json.dumps(layout_actions),
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "Org chart manual connector layout save failed")
        connection = sqlite3.connect(db_path)
        connector_count = connection.execute(
            """
            SELECT COUNT(*)
            FROM account_org_chart_connectors
            WHERE chart_id = ? AND orientation = ? AND source_side = ? AND target_side = ?
            """,
            (chart_id, "horizontal", "right", "left"),
        ).fetchone()[0]
        connection.close()
        assert_ok(connector_count == 1, "Org chart manual connector was not persisted with side-centre anchors")

        image = pipeflow_app.Image.new("RGB", (900, 320), (40, 120, 220))
        image_buffer = io.BytesIO()
        image.save(image_buffer, format="BMP")
        image_buffer.seek(0)
        image_upload = type("Upload", (), {"stream": image_buffer})()
        image_uri = pipeflow_app.upload_image_data_uri(image_upload, max_size=(160, 120))
        assert_ok(image_uri.startswith("data:image/png;base64,"), "Uploaded image was not normalised to PNG")
        normalised = pipeflow_app.Image.open(io.BytesIO(base64.b64decode(image_uri.split(",", 1)[1])))
        assert_ok(normalised.width <= 160 and normalised.height <= 120, "Uploaded image was not resized to fit display bounds")

        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        pg_context = pipeflow_app.pg_dashboard_context(connection)
        connection.close()
        plan_row = next(row for row in pg_context["pg_plan_rows"] if row["account_id"] == account_id)
        action_rows = [row for row in pg_context["pg_action_rows"] if row["account_id"] == account_id]
        assert_ok(action_rows, "PG Progress action rows missing for smoke account")
        assert_ok(
            any(row.get("contact_id") == second_contact_id for row in action_rows),
            "PG Progress did not display a contact with associated older outreach activity",
        )
        pg_progress_text = " ".join(
            str(entry)
            for row in action_rows
            for entry in (
                row.get("last_7_days_activity_entries", [])
                + row.get("next_7_days_actions", [])
            )
        )
        assert_ok(
            "Deleted PG Progress Marker" not in pg_progress_text,
            "PG Progress displayed a deleted outreach activity",
        )
        assert_ok(
            all(row.get("account_rag_status") == plan_row["rag_status"] for row in action_rows),
            "PG Progress lower account RAG does not match top account RAG",
        )
        assert_ok(plan_row["rag_status"] == "green", "PG Progress account RAG should reflect booked scheduled meeting evidence")
        pg_progress_html = client.get("/pg-progress").get_data(as_text=True)
        assert_ok("data-rag-trigger" in pg_progress_html, "PG Progress manual RAG picker trigger missing")
        response = client.post(
            "/pg-progress",
            data={
                "csrf_token": csrf_from_session(client),
                "current_pipeline": "1000",
                "pg_plan_account_id": [str(account_id)],
                f"rag_account_{account_id}": "green",
                "pg_action_contact_id": [str(contact_id)],
                f"pg_action_account_id_{contact_id}": str(account_id),
                f"rag_contact_{contact_id}": "amber",
                f"completed_discovery_contact_{contact_id}": "Yes",
                f"exec_first_contact_{contact_id}": "Yes",
                f"nbm_completed_contact_{contact_id}": "No",
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "PG Progress manual RAG save failed")
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        pg_context = pipeflow_app.pg_dashboard_context(connection)
        connection.close()
        plan_row = next(row for row in pg_context["pg_plan_rows"] if row["account_id"] == account_id)
        contact_row = next(row for row in pg_context["pg_action_rows"] if row["contact_id"] == contact_id)
        assert_ok(plan_row["rag_status"] == "green", "PG Progress account RAG did not persist manual selection")
        assert_ok(contact_row["rag_status"] == "amber", "PG Progress contact RAG did not persist manual selection")
        assert_ok(contact_row["account_rag_status"] == "green", "PG Progress action account RAG did not mirror manual account selection")
        response = client.post(
            "/pg-progress",
            data={
                "csrf_token": csrf_from_session(client),
                "current_pipeline": "1000",
                "pg_plan_account_id": [str(account_id)],
                f"rag_account_{account_id}": "",
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "PG Progress automatic RAG reset save failed")
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        pg_context = pipeflow_app.pg_dashboard_context(connection)
        connection.close()
        plan_row = next(row for row in pg_context["pg_plan_rows"] if row["account_id"] == account_id)
        assert_ok(plan_row["rag_status"] == "green", "PG Progress automatic RAG did not recalculate booked scheduled meeting evidence after manual override removal")
        assert_ok(plan_row["manual_rag_override"] == "", "PG Progress manual RAG override was not cleared")

        campaign_builder_html = client.get("/outreach/campaign-builder").get_data(as_text=True)
        assert_ok(
            "Smoke Test Play" in campaign_builder_html,
            "Campaign Builder sales play options missing on first load",
        )
        app_source = Path(pipeflow_app.__file__).read_text()
        assert_ok(
            "GROUP BY accounts.id\n        HAVING COUNT(contacts.id) > 0" not in app_source,
            "Campaign Builder contains a Postgres-incompatible account grouping query",
        )
        full_day = date(2026, 7, 22)
        full_day_reserved = {
            (full_day.isoformat(), f"{hour:02d}:{minute:02d}")
            for hour in range(9, 18)
            for minute in (0, 15, 30, 45)
            if not (hour == 17 and minute > 0)
        }
        direct_schedule = pipeflow_app.build_campaign_schedule(
            full_day,
            full_day + timedelta(days=2),
            1,
            1,
            reserved_slots=full_day_reserved,
        )
        assert_ok(
            direct_schedule[0]["action_date"] == full_day + timedelta(days=1),
            "Campaign Builder did not move to the next working day when the first day had no slots",
        )
        off_grid_reserved = {(full_day.isoformat(), "09:01")}
        direct_schedule = pipeflow_app.build_campaign_schedule(
            full_day,
            full_day,
            1,
            1,
            reserved_slots=off_grid_reserved,
        )
        assert_ok(
            direct_schedule[0]["time"] == "09:30",
            "Campaign Builder did not keep a 15-minute buffer from an off-grid scheduled event",
        )
        connection = sqlite3.connect(db_path)
        campaign_count_before = connection.execute(
            "SELECT COUNT(*) FROM outreach WHERE campaign = ?",
            ("Smoke Test Play",),
        ).fetchone()[0]
        connection.close()
        campaign_start = (today + timedelta(days=3)).isoformat()
        campaign_end = (today + timedelta(days=14)).isoformat()
        pg_week_start = (today + timedelta(days=21)).isoformat()
        response = client.post(
            "/outreach/campaign-builder",
            data={
                "csrf_token": csrf_from_session(client),
                "account_id": str(account_id),
                "pg_week_start": pg_week_start,
                "campaign_start_date": (today - timedelta(days=1)).isoformat(),
                "campaign_end_date": campaign_end,
                "total_outreach_tasks": "1",
                "times_per_week": "1",
                "sales_play": "Smoke Test Play",
                "fy": "27",
                "quarter": "Q1",
                "assigned_to": "Smoke Test Admin",
                "contact_ids": [str(contact_id)],
            },
            follow_redirects=True,
        )
        assert_ok(
            response.status_code == 200 and "Campaign Start cannot be earlier than" in response.get_data(as_text=True),
            "Campaign Builder did not reject a past campaign start date",
        )
        connection = sqlite3.connect(db_path)
        connection.execute("ALTER TABLE contacts DROP COLUMN personal_win")
        connection.execute("ALTER TABLE outreach DROP COLUMN campaign_start_date")
        connection.commit()
        connection.close()
        response = client.post(
            "/outreach/campaign-builder",
            data={
                "csrf_token": csrf_from_session(client),
                "account_id": str(account_id),
                "pg_week_start": pg_week_start,
                "campaign_start_date": campaign_start,
                "campaign_end_date": campaign_end,
                "total_outreach_tasks": "1",
                "times_per_week": "1",
                "sales_play": "Smoke Test Play",
                "fy": "27",
                "quarter": "Q1",
                "assigned_to": "Smoke Test Admin",
                "contact_ids": [str(contact_id)],
            },
            follow_redirects=False,
        )
        assert_ok(
            response.status_code in (302, 303)
            and response.headers.get("Location", "").endswith("/outreach"),
            "Campaign Builder success redirect should use a short Outreach URL without message query strings",
        )
        response = client.post(
            "/outreach/campaign-builder",
            data={
                "csrf_token": csrf_from_session(client),
                "account_id": str(account_id),
                "pg_week_start": pg_week_start,
                "campaign_start_date": campaign_start,
                "campaign_end_date": campaign_end,
                "total_outreach_tasks": "3",
                "times_per_week": "2",
                "sales_play": "Smoke Test Play",
                "fy": "27",
                "quarter": "Q1",
                "assigned_to": "Smoke Test Admin",
                "contact_ids": [str(contact_id)],
            },
            follow_redirects=True,
        )
        campaign_html = response.get_data(as_text=True)
        assert_ok(
            response.status_code == 200
            and "Outreach Tasks" in campaign_html
            and "Campaign created:" in campaign_html
            and "Smoke Test Contact" in campaign_html,
            "Campaign Builder did not redirect to Outreach with generated campaign confirmation",
        )
        connection = sqlite3.connect(db_path)
        campaign_count_after = connection.execute(
            "SELECT COUNT(*) FROM outreach WHERE campaign = ?",
            ("Smoke Test Play",),
        ).fetchone()[0]
        campaign_recipient_count = connection.execute(
            """
            SELECT COUNT(*)
            FROM outreach_recipients
            WHERE outreach_id IN (
                SELECT id
                FROM outreach
                WHERE campaign = ?
            )
            """,
            ("Smoke Test Play",),
        ).fetchone()[0]
        connection.close()
        assert_ok(campaign_count_after > campaign_count_before, "Campaign Builder did not save generated outreach")
        assert_ok(campaign_recipient_count > 0, "Campaign Builder did not save outreach recipients")

        new_campaign_contact_id = None
        new_contact_campaign_start = (today + timedelta(days=24)).isoformat()
        new_contact_campaign_end = (today + timedelta(days=34)).isoformat()
        connection = sqlite3.connect(db_path)
        new_campaign_contact_id = connection.execute(
            """
            INSERT INTO contacts (account_id, category, name, job_title, email, phone, location, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                account_id,
                "Executive",
                "Smoke Newly Added Contact",
                "VP Operations",
                "new-campaign-contact@example.com",
                "77777",
                "London",
                "Active",
            ),
        ).lastrowid
        connection.commit()
        connection.close()
        response = client.post(
            "/outreach/campaign-builder",
            data={
                "csrf_token": csrf_from_session(client),
                "account_id": str(account_id),
                "pg_week_start": (today + timedelta(days=40)).isoformat(),
                "campaign_start_date": new_contact_campaign_start,
                "campaign_end_date": new_contact_campaign_end,
                "total_outreach_tasks": "2",
                "times_per_week": "1",
                "sales_play": "Smoke Test Play",
                "fy": "27",
                "quarter": "Q1",
                "assigned_to": "Smoke Test Admin",
                "contact_ids": [str(new_campaign_contact_id)],
            },
            follow_redirects=True,
        )
        new_contact_campaign_html = response.get_data(as_text=True)
        assert_ok(
            response.status_code == 200
            and "Outreach Tasks" in new_contact_campaign_html
            and "Campaign created:" in new_contact_campaign_html
            and "Smoke Newly Added Contact" in new_contact_campaign_html,
            "Campaign Builder failed for a newly added contact on an existing account",
        )
        connection = sqlite3.connect(db_path)
        new_contact_campaign_count = connection.execute(
            """
            SELECT COUNT(*)
            FROM outreach
            WHERE contact_id = ?
              AND campaign = ?
              AND campaign_start_date = ?
            """,
            (new_campaign_contact_id, "Smoke Test Play", new_contact_campaign_start),
        ).fetchone()[0]
        connection.close()
        assert_ok(new_contact_campaign_count == 2, "Campaign Builder did not create the requested tasks for the newly added contact")

        response = client.post(
            "/outreach/campaign-builder",
            data={
                "csrf_token": csrf_from_session(client),
                "account_id": str(account_id),
                "pg_week_start": pg_week_start,
                "campaign_start_date": campaign_start,
                "campaign_end_date": campaign_end,
                "total_outreach_tasks": "2",
                "times_per_week": "1",
                "sales_play": "Smoke Test Play",
                "fy": "27",
                "quarter": "Q1",
                "assigned_to": "Smoke Test Admin",
            },
            follow_redirects=True,
        )
        assert_ok(
            response.status_code == 200
            and "Select at least one contact" in response.get_data(as_text=True),
            "Campaign Builder did not validate missing contacts clearly",
        )

        response = client.post(
            "/outreach/campaign-builder",
            data={
                "csrf_token": csrf_from_session(client),
                "account_id": str(account_id),
                "pg_week_start": pg_week_start,
                "campaign_start_date": campaign_start,
                "campaign_end_date": campaign_end,
                "total_outreach_tasks": "2",
                "times_per_week": "1",
                "sales_play": "Not Associated Play",
                "fy": "27",
                "quarter": "Q1",
                "assigned_to": "Smoke Test Admin",
                "contact_ids": [str(contact_id)],
            },
            follow_redirects=True,
        )
        assert_ok(
            response.status_code == 200
            and "Select a Sales Play used for or associated to the selected account." in response.get_data(as_text=True),
            "Campaign Builder did not validate account-specific Sales Play selection",
        )

        response = client.post(
            "/outreach/campaign-builder",
            data={
                "csrf_token": csrf_from_session(client),
                "account_id": str(account_id),
                "pg_week_start": pg_week_start,
                "campaign_start_date": campaign_end,
                "campaign_end_date": campaign_start,
                "total_outreach_tasks": "2",
                "times_per_week": "1",
                "sales_play": "Smoke Test Play",
                "fy": "27",
                "quarter": "Q1",
                "assigned_to": "Smoke Test Admin",
                "contact_ids": [str(contact_id)],
            },
            follow_redirects=True,
        )
        assert_ok(
            response.status_code == 200
            and "Campaign End cannot be earlier than Campaign Start" in response.get_data(as_text=True),
            "Campaign Builder did not validate reversed campaign dates clearly",
        )

        response = client.post(
            "/outreach/campaign-builder",
            data={
                "csrf_token": csrf_from_session(client),
                "account_id": str(account_id),
                "pg_week_start": pg_week_start,
                "campaign_start_date": campaign_start,
                "campaign_end_date": campaign_end,
                "total_outreach_tasks": "2",
                "times_per_week": "1",
                "sales_play": "Smoke Test Play",
                "fy": "27",
                "quarter": "Q1",
                "assigned_to": "Smoke Test Admin",
                "contact_ids": ["not-a-contact"],
            },
            follow_redirects=True,
        )
        assert_ok(
            response.status_code == 200
            and "One or more selected contacts are invalid" in response.get_data(as_text=True),
            "Campaign Builder did not validate malformed contact selections clearly",
        )

        filter_campaign_start = (today + timedelta(days=35)).isoformat()
        filter_campaign_end = (today + timedelta(days=45)).isoformat()
        connection = sqlite3.connect(db_path)
        connection.execute(
            """
            INSERT INTO outreach (
                fy, quarter, account_id, contact_id, campaign, sales_play,
                activity_date, activity_time, activity_type, subject, notes, outcome,
                next_action_date, next_action_time, task_status, assigned_to
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "27",
                "Q1",
                account_id,
                second_contact_id,
                "Existing",
                "Smoke Test Play",
                filter_campaign_start,
                "09:01",
                "Email",
                "Existing off-grid slot",
                "Blocks the early campaign slots.",
                "No Response",
                filter_campaign_start,
                "09:01",
                "Not Started",
                "Smoke Test Admin",
            ),
        )
        connection.execute("ALTER TABLE outreach_recipients DROP COLUMN sort_order")
        connection.commit()
        connection.close()
        response = client.post(
            "/outreach/campaign-builder",
            data={
                "csrf_token": csrf_from_session(client),
                "account_id": str(account_id),
                "pg_week_start": (today + timedelta(days=50)).isoformat(),
                "campaign_start_date": filter_campaign_start,
                "campaign_end_date": filter_campaign_end,
                "total_outreach_tasks": "3",
                "times_per_week": "2",
                "sales_play": "Smoke Test Play",
                "fy": "27",
                "quarter": "Q1",
                "assigned_to": "Smoke Test Admin",
                "contact_ids": [str(second_contact_id)],
                "campaign_activity_types": ["Phone"],
            },
            follow_redirects=True,
        )
        filtered_campaign_html = response.get_data(as_text=True)
        assert_ok(
            response.status_code == 200
            and "Outreach Tasks" in filtered_campaign_html
            and "Campaign created:" in filtered_campaign_html,
            "Campaign Builder did not redirect to Outreach after selected activity generation and recipient schema recovery",
        )
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        filtered_types = [
            row["activity_type"]
            for row in connection.execute(
                """
                SELECT activity_type
                FROM outreach
                WHERE campaign = ?
                  AND contact_id = ?
                  AND campaign_start_date = ?
                ORDER BY next_action_date, next_action_time
                """,
                ("Smoke Test Play", second_contact_id, filter_campaign_start),
            ).fetchall()
        ]
        filtered_times = [
            row["next_action_time"]
            for row in connection.execute(
                """
                SELECT next_action_time
                FROM outreach
                WHERE campaign = ?
                  AND contact_id = ?
                  AND campaign_start_date = ?
                ORDER BY next_action_date, next_action_time
                """,
                ("Smoke Test Play", second_contact_id, filter_campaign_start),
            ).fetchall()
        ]
        filtered_recipient_count = connection.execute(
            """
            SELECT COUNT(*)
            FROM outreach_recipients
            WHERE outreach_id IN (
                SELECT id
                FROM outreach
                WHERE campaign = ?
                  AND contact_id = ?
                  AND campaign_start_date = ?
            )
            """,
            ("Smoke Test Play", second_contact_id, filter_campaign_start),
        ).fetchone()[0]
        connection.close()
        assert_ok(filtered_types and filtered_types[0] == "VITO", "Campaign Builder did not keep VITO as the first filtered campaign step")
        assert_ok(set(filtered_types[1:]).issubset({"Phone"}), "Campaign Builder generated an unselected activity type after VITO")
        assert_ok(
            all(int(time_value.split(":")[1]) % 15 == 0 for time_value in filtered_times),
            "Campaign Builder generated a task outside the 15-minute scheduling grid",
        )
        assert_ok(
            "09:00" not in filtered_times and "09:15" not in filtered_times,
            "Campaign Builder scheduled inside the 15-minute buffer around an existing off-grid task",
        )
        assert_ok(filtered_recipient_count == len(filtered_types), "Campaign Builder did not recover and save recipient links")

        response = client.post(
            "/outreach/campaign-builder",
            data={
                "csrf_token": csrf_from_session(client),
                "account_id": str(account_id),
                "pg_week_start": (today + timedelta(days=50)).isoformat(),
                "campaign_start_date": filter_campaign_start,
                "campaign_end_date": filter_campaign_end,
                "total_outreach_tasks": "3",
                "times_per_week": "2",
                "sales_play": "Smoke Test Play",
                "fy": "27",
                "quarter": "Q1",
                "assigned_to": "Smoke Test Admin",
                "contact_ids": [str(second_contact_id)],
                "campaign_activity_types": ["Phone"],
            },
            follow_redirects=True,
        )
        repeated_campaign_html = response.get_data(as_text=True)
        assert_ok(
            response.status_code == 200
            and "Outreach Tasks" in repeated_campaign_html
            and "This campaign has already been generated" in repeated_campaign_html
            and "same campaign step already exists" not in repeated_campaign_html
            and "Internal Server Error" not in repeated_campaign_html,
            "Campaign Builder did not handle repeated campaign generation as a safe no-op",
        )
        assert_ok(
            "same campaign step already exists" in app_source,
            "Campaign Builder duplicate warning does not explain exact campaign duplicate criteria",
        )

        original_store_page_notice = pipeflow_app.store_page_notice
        notice_fallback_start = (today + timedelta(days=62)).isoformat()
        notice_fallback_end = (today + timedelta(days=66)).isoformat()
        pipeflow_app.store_page_notice = lambda message="", error="": False
        try:
            response = client.post(
                "/outreach/campaign-builder",
                data={
                    "csrf_token": csrf_from_session(client),
                    "account_id": str(account_id),
                    "pg_week_start": (today + timedelta(days=72)).isoformat(),
                    "campaign_start_date": notice_fallback_start,
                    "campaign_end_date": notice_fallback_end,
                    "total_outreach_tasks": "1",
                    "times_per_week": "1",
                    "sales_play": "Smoke Test Play",
                    "fy": "27",
                    "quarter": "Q1",
                    "assigned_to": "Smoke Test Admin",
                    "contact_ids": [str(second_contact_id)],
                },
                follow_redirects=True,
            )
        finally:
            pipeflow_app.store_page_notice = original_store_page_notice
        notice_fallback_html = response.get_data(as_text=True)
        assert_ok(
            response.status_code == 200
            and "Outreach Tasks" in notice_fallback_html
            and "Campaign created:" in notice_fallback_html
            and "Internal Server Error" not in notice_fallback_html,
            "Campaign Builder notice fallback did not redirect safely after creating tasks",
        )
        connection = sqlite3.connect(db_path)
        notice_fallback_count = connection.execute(
            """
            SELECT COUNT(*)
            FROM outreach
            WHERE campaign = ?
              AND contact_id = ?
              AND campaign_start_date = ?
            """,
            ("Smoke Test Play", second_contact_id, notice_fallback_start),
        ).fetchone()[0]
        connection.close()
        assert_ok(notice_fallback_count == 1, "Campaign Builder notice fallback did not save the generated task")

        original_campaign_builder_impl = pipeflow_app.campaign_builder_impl

        def broken_campaign_builder_impl():
            raise RuntimeError("forced campaign builder smoke failure")

        pipeflow_app.campaign_builder_impl = broken_campaign_builder_impl
        try:
            response = client.post(
                "/outreach/campaign-builder",
                data={
                    "csrf_token": csrf_from_session(client),
                    "account_id": str(account_id),
                    "pg_week_start": (today + timedelta(days=50)).isoformat(),
                    "campaign_start_date": filter_campaign_start,
                    "campaign_end_date": filter_campaign_end,
                    "total_outreach_tasks": "3",
                    "times_per_week": "2",
                    "sales_play": "Smoke Test Play",
                    "fy": "27",
                    "quarter": "Q1",
                    "assigned_to": "Smoke Test Admin",
                    "contact_ids": [str(second_contact_id)],
                },
                follow_redirects=True,
            )
        finally:
            pipeflow_app.campaign_builder_impl = original_campaign_builder_impl
        failed_campaign_html = response.get_data(as_text=True)
        assert_ok(
            response.status_code == 200
            and "Outreach Tasks" in failed_campaign_html
            and "Campaign could not be generated" in failed_campaign_html
            and "Internal Server Error" not in failed_campaign_html,
            "Campaign Builder POST failure did not redirect back to Outreach with a human-readable error",
        )
        original_outreach_impl = pipeflow_app.outreach_impl

        def broken_outreach_impl():
            raise RuntimeError("forced outreach smoke failure")

        pipeflow_app.outreach_impl = broken_outreach_impl
        try:
            response = client.get("/outreach?message=Campaign%20created")
        finally:
            pipeflow_app.outreach_impl = original_outreach_impl
        failed_outreach_html = response.get_data(as_text=True)
        assert_ok(
            response.status_code == 200
            and "Internal Server Error" not in failed_outreach_html
            and "Outreach could not load the table" in failed_outreach_html,
            "Outreach failure fallback did not show a safe human-readable page",
        )
        connection = sqlite3.connect(db_path)
        duplicate_campaign_rows = connection.execute(
            """
            SELECT COUNT(*)
            FROM (
                SELECT contact_id, next_action_date, next_action_time, activity_type, subject, COUNT(*) AS total
                FROM outreach
                WHERE campaign = ?
                  AND contact_id = ?
                  AND campaign_start_date = ?
                GROUP BY contact_id, next_action_date, next_action_time, activity_type, subject
                HAVING COUNT(*) > 1
            )
            """,
            ("Smoke Test Play", second_contact_id, filter_campaign_start),
        ).fetchone()[0]
        connection.close()
        assert_ok(
            duplicate_campaign_rows == 0,
            "Campaign Builder created duplicate contact/date/time/activity/subject rows",
        )

        false_duplicate_start = (today + timedelta(days=70)).isoformat()
        false_duplicate_end = (today + timedelta(days=75)).isoformat()
        connection = sqlite3.connect(db_path)
        connection.execute(
            """
            INSERT INTO outreach (
                fy, quarter, account_id, contact_id, campaign, sales_play,
                activity_date, activity_time, activity_type, subject, notes, outcome,
                next_action_date, next_action_time, task_status, assigned_to
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "27",
                "Q1",
                account_id,
                second_contact_id,
                "",
                "Smoke Test Play",
                false_duplicate_start,
                "09:00",
                "VITO",
                "VITO: Smoke Test Play",
                "Existing single outreach should not block first campaign generation.",
                "No Response",
                false_duplicate_start,
                "09:00",
                "Not Started",
                "Smoke Test Admin",
            ),
        )
        connection.commit()
        connection.close()
        response = client.post(
            "/outreach/campaign-builder",
            data={
                "csrf_token": csrf_from_session(client),
                "account_id": str(account_id),
                "pg_week_start": (today + timedelta(days=80)).isoformat(),
                "campaign_start_date": false_duplicate_start,
                "campaign_end_date": false_duplicate_end,
                "total_outreach_tasks": "1",
                "times_per_week": "1",
                "sales_play": "Smoke Test Play",
                "fy": "27",
                "quarter": "Q1",
                "assigned_to": "Smoke Test Admin",
                "contact_ids": [str(second_contact_id)],
            },
            follow_redirects=True,
        )
        false_duplicate_html = response.get_data(as_text=True)
        assert_ok(
            response.status_code == 200
            and "Campaign created:" in false_duplicate_html
            and "same campaign step already exists" not in false_duplicate_html,
            "Campaign Builder falsely treated an existing single outreach task as a duplicate campaign step",
        )

        response = client.post(
            "/outreach/campaign-builder",
            data={
                "csrf_token": csrf_from_session(client),
                "account_id": str(account_id),
                "pg_week_start": pg_week_start,
                "campaign_start_date": campaign_start,
                "campaign_end_date": campaign_end,
                "total_outreach_tasks": "not-a-number",
                "times_per_week": "2",
                "sales_play": "Smoke Test Play",
                "fy": "27",
                "quarter": "Q1",
                "assigned_to": "Smoke Test Admin",
                "contact_ids": [str(contact_id)],
            },
            follow_redirects=True,
        )
        assert_ok(
            response.status_code == 200 and "Qty Outreach Tasks must be a whole number from 1 to 50." in response.get_data(as_text=True),
            "Campaign Builder did not show field-specific quantity validation",
        )

        dashboard_html = client.get("/").get_data(as_text=True)
        assert_ok("header-action-stack" in dashboard_html, "header action stack missing")
        assert_ok(
            dashboard_html.index("User Guide") < dashboard_html.index("Release Notes") < dashboard_html.index("<nav"),
            "Release Notes is not stacked below User Guide before the main nav",
        )

        add_account_html = client.get("/accounts/add").get_data(as_text=True)
        assert_ok(
            'name="pg_bible_order"' not in add_account_html
            and "assigns the next available PG Bible number" in add_account_html,
            "Account creation still allows manual PG Bible Order entry",
        )
        connection = sqlite3.connect(db_path)
        expected_pg_bible_order = connection.execute(
            "SELECT COALESCE(MAX(pg_bible_order), 0) + 1 FROM accounts"
        ).fetchone()[0]
        connection.close()
        response = client.post(
            "/accounts/add",
            data={
                "csrf_token": csrf_from_session(client),
                "account_name": "Smoke Auto Number Account",
                "country": "United Kingdom",
                "account_tier": "2",
                "pipeline_target": "50000",
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "Account auto-number creation failed")
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        auto_number_account = connection.execute(
            "SELECT id, pg_bible_order FROM accounts WHERE account_name = ?",
            ("Smoke Auto Number Account",),
        ).fetchone()
        connection.close()
        assert_ok(
            auto_number_account is not None
            and auto_number_account["pg_bible_order"] == expected_pg_bible_order,
            "Account creation did not assign the next PG Bible Order",
        )
        response = client.post(
            f"/accounts/{auto_number_account['id']}/edit",
            data={
                "csrf_token": csrf_from_session(client),
                "account_name": "Smoke Auto Number Account",
                "pg_bible_order": "1",
                "country": "United Kingdom",
                "account_tier": "2",
                "pipeline_target": "50000",
            },
            follow_redirects=False,
        )
        assert_ok(
            response.status_code == 200
            and "is already allocated to another account" in response.get_data(as_text=True),
            "Duplicate PG Bible Order edit did not return clear validation",
        )
        response = client.post(
            f"/accounts/{auto_number_account['id']}/delete",
            data={"csrf_token": csrf_from_session(client)},
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "Smoke auto-number account cleanup failed")

        response = client.post(
            f"/tasks/{outreach_id}/update",
            data={
                "csrf_token": csrf_from_session(client),
                "return_to": "/",
                "next_action": "Updated smoke follow up",
                "next_action_date": "2026-08-06",
                "next_action_time": "11:30",
                "task_status": "In Progress",
                "outcome": "No Response",
                "notes": "Updated from smoke test",
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "dashboard task update failed")

        response = client.get("/outreach/add")
        add_html = response.get_data(as_text=True)
        for outcome in ("NBM Booked", "Discovery Booked", "Exec Meeting Booked"):
            assert_ok(outcome in add_html, f"{outcome} outcome missing from Outreach form")
        assert_ok("Close and Create New" in add_html, "Add Outreach close-and-new button missing")
        assert_ok(
            "Auto Schedule" in add_html and "outreach-auto-schedule-panel" in add_html,
            "Add Outreach auto-schedule button missing",
        )
        assert_ok(
            "full-width outreach-auto-schedule-panel" not in add_html,
            "Add Outreach auto-schedule control should not be full width",
        )
        for field_id in ("activity_date", "activity_time", "next_action_date", "next_action_time", "scheduled_meeting_at"):
            assert_ok(input_value(add_html, field_id) == "", f"new Outreach {field_id} should open blank")
        for prefill_url in (
            f"/outreach/add?account_id={account_id}",
            f"/outreach/add?contact_id={contact_id}",
            f"/outreach/add?prefill_from={outreach_id}",
        ):
            prefilled_add_html = client.get(prefill_url).get_data(as_text=True)
            for field_id in ("activity_date", "activity_time", "next_action_date", "next_action_time"):
                assert_ok(
                    input_value(prefilled_add_html, field_id) == "",
                    f"prefilled new Outreach {field_id} should remain blank",
                )
        activity_start_input = re.search(r'<input[^>]+id="activity_date"[^>]*>', add_html)
        due_date_input = re.search(r'<input[^>]+id="next_action_date"[^>]*>', add_html)
        assert_ok(
            activity_start_input and "min=" not in activity_start_input.group(0),
            "Activity Start Date should allow manual backdating",
        )
        assert_ok(
            due_date_input and "min=" not in due_date_input.group(0),
            "Activity Due Date should allow manual retrospective scheduling",
        )
        assert_ok(
            "Manual backdating is allowed" in add_html,
            "Activity Start backdating guidance missing",
        )

        campaign_builder_html = client.get("/outreach/campaign-builder").get_data(as_text=True)
        for field_id in ("campaign_start_date", "campaign_end_date", "pg_week_start"):
            assert_ok(input_value(campaign_builder_html, field_id) == "", f"Campaign Builder {field_id} should open blank")
        assert_ok("setDefaultCampaignWindow" not in campaign_builder_html, "Campaign Builder still auto-populates dates in the browser")
        prefilled_campaign_html = client.get(
            f"/outreach/campaign-builder?account_id={account_id}&contact_id={contact_id}"
        ).get_data(as_text=True)
        for field_id in ("campaign_start_date", "campaign_end_date", "pg_week_start"):
            assert_ok(
                input_value(prefilled_campaign_html, field_id) == "",
                f"prefilled Campaign Builder {field_id} should remain blank",
            )

        response = client.get(f"/outreach/{outreach_id}/edit")
        edit_outreach_html = response.get_data(as_text=True)
        assert_ok(
            response.status_code == 200
            and "Smoke Test Account - BMC" in edit_outreach_html,
            "Edit Outreach account selector does not show account business/org",
        )

        response = client.get(f"/outreach/add?account_id={account_id}")
        prefill_account_html = response.get_data(as_text=True)
        assert_ok(
            response.status_code == 200
            and f'value="{account_id}" selected' in prefill_account_html,
            "account quick-create Outreach prefill missing",
        )

        response = client.get(f"/outreach/add?account_id={account_id}&contact_id={contact_id}")
        prefill_contact_html = response.get_data(as_text=True)
        assert_ok(
            response.status_code == 200
            and f'value="{account_id}" selected' in prefill_contact_html
            and f'value="{contact_id}"' in prefill_contact_html
            and "checked" in prefill_contact_html,
            "contact quick-create Outreach prefill missing",
        )

        response = client.get("/admin/permissions")
        admin_html = response.get_data(as_text=True)
        assert_ok(
            response.status_code == 200
            and "broadcast-company-dropdown" in admin_html
            and "admin-company-multiselect" in admin_html,
            "admin company membership or broadcast controls missing",
        )
        assert_ok("target_all_companies" in admin_html, "global broadcast option missing")

        response = client.post(
            "/admin/tenants",
            data={
                "csrf_token": csrf_from_session(client),
                "company_name": "Smoke Other Company",
                "country": "United Kingdom",
                "company_contact": "Smoke Tenant Owner",
            },
            follow_redirects=True,
        )
        assert_ok(response.status_code == 200, "tenant create failed")

        response = client.post(
            "/admin/users/create",
            data={
                "csrf_token": csrf_from_session(client),
                "full_name": "Smoke Multi Company Admin",
                "email": "smoke-multi-admin@example.com",
                "password": "Password123!",
                "reset_phrase": "multi company phrase",
                "company": "PipeFlow Administration",
                "company_memberships": ["PipeFlow Administration", "Smoke Other Company"],
                "role": "admin",
                "team_ids": [],
            },
            follow_redirects=True,
        )
        assert_ok(response.status_code == 200, "multi-company admin create failed")

        response = client.post(
            "/admin/broadcasts/add",
            data={
                "csrf_token": csrf_from_session(client),
                "title": "Smoke broadcast",
                "message": "Smoke test message",
                "severity": "info",
                "target_companies": ["PipeFlow Administration", "Smoke Other Company"],
                "start_at": f"{yesterday}T09:00",
                "stop_at": f"{tomorrow}T09:00",
                "is_active": "1",
            },
            follow_redirects=True,
        )
        broadcast_create_html = response.get_data(as_text=True)
        assert_ok(response.status_code == 200 and "Smoke broadcast" in broadcast_create_html, "broadcast create failed")
        assert_ok(
            "Save Changes" in broadcast_create_html
            and "Pause Broadcast" in broadcast_create_html
            and "Delete Broadcast" in broadcast_create_html,
            "broadcast edit action buttons are not grouped on existing broadcasts",
        )

        import auth as pipeflow_auth

        auth_connection = pipeflow_auth.get_auth_connection()
        multi_admin = auth_connection.execute(
            "SELECT id FROM users WHERE email = ?",
            ("smoke-multi-admin@example.com",),
        ).fetchone()
        smoke_broadcast = auth_connection.execute(
            "SELECT id, target_companies FROM broadcast_messages WHERE title = ?",
            ("Smoke broadcast",),
        ).fetchone()
        auth_connection.close()
        assert_ok(multi_admin is not None, "multi-company admin user not found")
        assert_ok(
            set(pipeflow_auth.user_company_names(multi_admin["id"], include_primary=True)) >= {"PipeFlow Administration", "Smoke Other Company"},
            "multi-company admin memberships were not saved",
        )
        assert_ok(smoke_broadcast is not None, "created broadcast not found")

        response = client.post(
            f"/admin/broadcasts/{smoke_broadcast['id']}/update",
            data={
                "csrf_token": csrf_from_session(client),
                "title": "Smoke broadcast updated",
                "message": "Smoke test message updated",
                "severity": "warning",
                "target_companies": ["Smoke Other Company"],
                "start_at": f"{yesterday}T10:00",
                "stop_at": f"{tomorrow}T10:00",
                "is_active": "1",
            },
            follow_redirects=True,
        )
        assert_ok(response.status_code == 200 and "Smoke broadcast updated" in response.get_data(as_text=True), "broadcast update failed")

        updated_broadcast = pipeflow_auth.get_broadcast_message(smoke_broadcast["id"])
        assert_ok(
            pipeflow_auth.decode_broadcast_companies(updated_broadcast["target_companies"]) == ["Smoke Other Company"],
            "broadcast company targets were not updated",
        )

        response = client.post(
            "/admin/broadcasts/add",
            data={
                "csrf_token": csrf_from_session(client),
                "title": "Smoke global broadcast",
                "message": "Smoke test message for every company",
                "severity": "info",
                "target_all_companies": "1",
                "start_at": f"{yesterday}T11:00",
                "stop_at": f"{tomorrow}T11:00",
                "is_active": "1",
            },
            follow_redirects=True,
        )
        assert_ok(response.status_code == 200 and "Smoke global broadcast" in response.get_data(as_text=True), "global broadcast create failed")
        auth_connection = pipeflow_auth.get_auth_connection()
        global_broadcast = auth_connection.execute(
            "SELECT target_companies FROM broadcast_messages WHERE title = ?",
            ("Smoke global broadcast",),
        ).fetchone()
        auth_connection.close()
        assert_ok(
            global_broadcast is not None
            and pipeflow_auth.decode_broadcast_companies(global_broadcast["target_companies"]) == [],
            "global broadcast should have no company target filter",
        )

        manual_past_due_date = (today - timedelta(days=1)).isoformat()
        response = client.post(
            "/outreach/add",
            data={
                "csrf_token": csrf_from_session(client),
                "fy": "27",
                "quarter": "Q1",
                "account_id": str(account_id),
                "sales_play": "Smoke Test Play",
                "contact_ids": [str(contact_id)],
                "task_status": "Not Started",
                "assigned_to": "Smoke Test Admin",
                "activity_type": "Call",
                "activity_date": "2026-05-07",
                "activity_time": "09:00",
                "next_action_date": manual_past_due_date,
                "next_action_time": "10:00",
                "subject": "Smoke allowed manual past due outreach",
                "outcome": "No Response",
                "next_action": "",
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "new Outreach manually entered past due date was blocked")
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        manual_past_due_outreach = connection.execute(
            "SELECT next_action_date, next_action_time FROM outreach WHERE subject = ?",
            ("Smoke allowed manual past due outreach",),
        ).fetchone()
        connection.close()
        assert_ok(
            manual_past_due_outreach
            and manual_past_due_outreach["next_action_date"] == manual_past_due_date
            and manual_past_due_outreach["next_action_time"] == "10:00",
            "new Outreach manually entered past due date was not saved",
        )

        backdated_activity_date = (today - timedelta(days=30)).isoformat()
        response = client.post(
            "/outreach/add",
            data={
                "csrf_token": csrf_from_session(client),
                "fy": "27",
                "quarter": "Q1",
                "account_id": str(account_id),
                "sales_play": "Smoke Test Play",
                "contact_ids": [str(second_contact_id)],
                "task_status": "Not Started",
                "assigned_to": "Smoke Test Admin",
                "activity_type": "Email",
                "activity_date": backdated_activity_date,
                "activity_time": "08:00",
                "next_action_date": (today + timedelta(days=12)).isoformat(),
                "next_action_time": "10:00",
                "subject": "Smoke allowed backdated activity start",
                "outcome": "No Response",
                "next_action": "",
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "manual backdated Activity Start was blocked on new Outreach")
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        backdated_outreach = connection.execute(
            "SELECT activity_date, activity_time FROM outreach WHERE subject = ?",
            ("Smoke allowed backdated activity start",),
        ).fetchone()
        connection.close()
        assert_ok(
            backdated_outreach
            and backdated_outreach["activity_date"] == backdated_activity_date
            and backdated_outreach["activity_time"] == "08:00",
            "manual backdated Activity Start was not saved as entered",
        )

        buffer_conflict_date = today
        connection = sqlite3.connect(db_path)
        connection.execute(
            """
            INSERT INTO outreach (
                fy, quarter, account_id, contact_id, campaign, sales_play,
                activity_date, activity_time, activity_type, subject, notes, outcome,
                next_action_date, next_action_time, task_status, assigned_to
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "27",
                "Q1",
                account_id,
                contact_id,
                "Smoke Test Play",
                "Smoke Test Play",
                buffer_conflict_date.isoformat(),
                "09:00",
                "Call",
                "Smoke auto schedule contact buffer source",
                "Blocks auto-schedule within two days.",
                "No Response",
                buffer_conflict_date.isoformat(),
                "09:00",
                "Not Started",
                "Smoke Test Admin",
            ),
        )
        connection.commit()
        connection.close()
        response = client.post(
            "/outreach/auto-schedule",
            data={
                "csrf_token": csrf_from_session(client),
                "account_id": str(account_id),
                "contact_ids": [str(contact_id)],
            },
        )
        auto_schedule_payload = response.get_json()
        assert_ok(response.status_code == 200 and auto_schedule_payload and auto_schedule_payload.get("ok"), "new Outreach auto-schedule did not return a slot")
        auto_scheduled_date = date.fromisoformat(auto_schedule_payload["activity_date"])
        assert_ok(auto_scheduled_date >= today, "new Outreach auto-schedule returned a past date")
        assert_ok(
            auto_schedule_payload.get("next_action_date") == auto_schedule_payload.get("activity_date")
            and auto_schedule_payload.get("next_action_time") == auto_schedule_payload.get("activity_time"),
            "new Outreach auto-schedule did not populate matching due/end fields",
        )
        assert_ok(abs((auto_scheduled_date - buffer_conflict_date).days) > 2, "new Outreach auto-schedule ignored the two-day contact buffer")

        response = client.post(
            "/outreach/add",
            data={
                "csrf_token": csrf_from_session(client),
                "fy": "27",
                "quarter": "Q1",
                "account_id": str(account_id),
                "sales_play": "Smoke Test Play",
                "contact_ids": [str(contact_id), str(second_contact_id)],
                "task_status": "Not Started",
                "assigned_to": "Smoke Test Admin",
                "activity_type": "Call",
                "activity_date": "2026-05-07",
                "activity_time": "09:00",
                "next_action_date": "2026-08-10",
                "next_action_time": "10:00",
                "subject": "Smoke multi-contact outreach",
                "outcome": "NBM Booked",
                "scheduled_meeting_at": f"{future_meeting_date}T09:30",
                "next_action": "",
            },
            follow_redirects=False,
        )
        assert_ok(
            response.status_code in (302, 303),
            f"multi-contact outreach add failed ({response.status_code}): {response.get_data(as_text=True)[:5000]}",
        )

        response = client.post(
            "/outreach/add",
            data={
                "csrf_token": csrf_from_session(client),
                "submit_action": "close_and_new",
                "fy": "27",
                "quarter": "Q1",
                "account_id": str(account_id),
                "sales_play": "Smoke Test Play",
                "contact_ids": [str(contact_id)],
                "task_status": "Not Started",
                "assigned_to": "Smoke Test Admin",
                "activity_type": "Email",
                "activity_date": "2026-05-07",
                "activity_time": "09:30",
                "next_action_date": "2026-08-11",
                "next_action_time": "11:00",
                "subject": "Smoke close and create new outreach",
                "outcome": "No Response",
                "next_action": "Closed immediately for reporting smoke coverage.",
            },
            follow_redirects=False,
        )
        close_new_location = response.headers.get("Location", "")
        assert_ok(
            response.status_code in (302, 303) and "/outreach/add?prefill_from=" in close_new_location,
            "Add Outreach Close and Create New did not redirect to a prefilled new Outreach form",
        )
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        close_new_outreach = connection.execute(
            "SELECT id, task_status, next_action_date, next_action_time FROM outreach WHERE subject = ?",
            ("Smoke close and create new outreach",),
        ).fetchone()
        connection.close()
        assert_ok(close_new_outreach is not None, "Add Outreach Close and Create New did not save the outreach")
        assert_ok(close_new_outreach["task_status"] == "Completed", "Add Outreach Close and Create New did not complete the saved outreach")
        assert_ok(not close_new_outreach["next_action_date"] and not close_new_outreach["next_action_time"], "Add Outreach Close and Create New should clear due date fields")

        response = client.post(
            "/outreach/add",
            data={
                "csrf_token": csrf_from_session(client),
                "fy": "27",
                "quarter": "Q1",
                "account_id": str(account_id),
                "sales_play": "Smoke Test Play",
                "contact_ids": [f"partner_contact:{partner_contact_id}"],
                "task_status": "Not Started",
                "assigned_to": "Smoke Test Admin",
                "activity_type": "Meeting",
                "activity_date": "2026-05-13",
                "activity_time": "09:00",
                "next_action_date": "2026-08-20",
                "next_action_time": "10:00",
                "subject": "Smoke account representative meeting outreach",
                "outcome": "Positive Response",
                "next_action": "",
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "account representative meeting Outreach add failed")
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        representative_outreach = connection.execute(
            "SELECT partner_contact_id FROM outreach WHERE subject = ?",
            ("Smoke account representative meeting outreach",),
        ).fetchone()
        connection.close()
        assert_ok(
            representative_outreach is not None
            and str(representative_outreach["partner_contact_id"]) == str(partner_contact_id),
            "account representative meeting Outreach was not saved",
        )

        response = client.post(
            "/outreach/add",
            data={
                "csrf_token": csrf_from_session(client),
                "fy": "27",
                "quarter": "Q1",
                "account_id": str(account_id),
                "sales_play": "Smoke Test Play",
                "contact_ids": ["partner_contact:not-a-number"],
                "task_status": "Not Started",
                "assigned_to": "Smoke Test Admin",
                "activity_type": "Meeting",
                "activity_date": "2026-05-13",
                "activity_time": "09:30",
                "next_action_date": "2026-08-20",
                "next_action_time": "10:30",
                "subject": "Smoke malformed representative outreach",
                "outcome": "Positive Response",
            },
            follow_redirects=True,
        )
        malformed_html = response.get_data(as_text=True)
        assert_ok(
            response.status_code == 200 and "Select a contact or partner contact" in malformed_html,
            "malformed representative id did not return validation",
        )

        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        other_account_id = connection.execute(
            """
            INSERT INTO accounts (
                account_name, pg_bible_order, account_tier, industry, business_unit,
                country, city, website, pipeline_target, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "Smoke Other Account",
                2,
                "2",
                "Technology",
                "BMC",
                "United Kingdom",
                "London",
                "https://other.example.com",
                250000,
                "Smoke test other account",
            ),
        ).lastrowid
        other_contact_id = connection.execute(
            """
            INSERT INTO contacts (account_id, category, name, job_title, email, phone, location)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (other_account_id, "Technical", "Smoke Other Contact", "Architect", "other@example.com", "55555", "London"),
        ).lastrowid
        connection.commit()
        multi_outreach = connection.execute(
            "SELECT * FROM outreach WHERE subject = ?",
            ("Smoke multi-contact outreach",),
        ).fetchone()
        assert_ok(multi_outreach is not None, "multi-contact outreach was not saved")
        assert_ok(str(multi_outreach["contact_id"]) == str(contact_id), "primary outreach contact was not preserved")
        assert_ok(multi_outreach["outcome"] == "NBM Booked", "NBM Booked outcome was not saved")
        recipient_count = connection.execute(
            "SELECT COUNT(*) FROM outreach_recipients WHERE outreach_id = ?",
            (multi_outreach["id"],),
        ).fetchone()[0]
        connection.close()
        assert_ok(recipient_count == 2, "multi-contact outreach recipients were not saved")

        connection = sqlite3.connect(db_path)
        connection.execute("DROP TABLE outreach_recipients")
        connection.commit()
        connection.close()
        response = client.post(
            "/outreach/add",
            data={
                "csrf_token": csrf_from_session(client),
                "fy": "27",
                "quarter": "Q1",
                "account_id": str(account_id),
                "sales_play": "Smoke Test Play",
                "contact_ids": [str(contact_id)],
                "task_status": "Not Started",
                "assigned_to": "Smoke Test Admin",
                "activity_type": "Email",
                "activity_date": "2026-05-09",
                "activity_time": "08:15",
                "next_action_date": "2026-08-15",
                "next_action_time": "09:15",
                "subject": "Smoke schema-refresh outreach",
                "outcome": "No Response",
                "next_action": "Retry after schema refresh",
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "schema-refresh Outreach add failed")
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        recovered_outreach = connection.execute(
            "SELECT id FROM outreach WHERE subject = ?",
            ("Smoke schema-refresh outreach",),
        ).fetchone()
        recovered_recipients = connection.execute(
            "SELECT COUNT(*) FROM outreach_recipients WHERE outreach_id = ?",
            (recovered_outreach["id"] if recovered_outreach else None,),
        ).fetchone()[0]
        connection.close()
        assert_ok(recovered_outreach is not None, "schema-refresh Outreach was not saved")
        assert_ok(recovered_recipients == 1, "schema-refresh Outreach recipient was not saved")

        connection = sqlite3.connect(db_path)
        connection.execute("DROP TABLE audit_entries")
        connection.commit()
        connection.close()
        response = client.post(
            "/outreach/add",
            data={
                "csrf_token": csrf_from_session(client),
                "fy": "27",
                "quarter": "Q1",
                "account_id": str(account_id),
                "sales_play": "Smoke Test Play",
                "contact_ids": [str(contact_id)],
                "task_status": "Not Started",
                "assigned_to": "Smoke Test Admin",
                "activity_type": "Email",
                "activity_date": "2026-05-09",
                "activity_time": "08:45",
                "next_action_date": "2026-08-16",
                "next_action_time": "09:45",
                "subject": "Smoke audit recovery outreach",
                "outcome": "No Response",
                "next_action": "Audit recovery should not block save",
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "audit recovery Outreach add failed")
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        audit_recovered_outreach = connection.execute(
            "SELECT id FROM outreach WHERE subject = ?",
            ("Smoke audit recovery outreach",),
        ).fetchone()
        connection.close()
        assert_ok(audit_recovered_outreach is not None, "audit recovery Outreach was not saved")

        response = client.post(
            "/outreach/add",
            data={
                "csrf_token": csrf_from_session(client),
                "fy": "27",
                "quarter": "Q1",
                "account_id": str(account_id),
                "sales_play": "Smoke Test Play",
                "contact_ids": [str(other_contact_id)],
                "task_status": "Not Started",
                "assigned_to": "Smoke Test Admin",
                "activity_type": "Call",
                "activity_date": "2026-05-07",
                "activity_time": "09:00",
                "next_action_date": "2026-08-10",
                "next_action_time": "10:00",
                "subject": "Smoke mismatched-contact outreach",
                "outcome": "No Response Yet",
                "next_action": "",
            },
            follow_redirects=True,
        )
        mismatch_html = response.get_data(as_text=True)
        assert_ok(
            response.status_code == 200
            and "Select a contact or partner contact that belongs to the selected account." in mismatch_html,
            "mismatched account/contact outreach was not rejected",
        )

        connection = sqlite3.connect(db_path)
        historical_due_outreach_id = connection.execute(
            """
            INSERT INTO outreach (
                fy, quarter, account_id, contact_id, campaign, sales_play,
                activity_date, activity_time, activity_type, subject, outcome,
                next_action, next_action_date, next_action_time, task_status, assigned_to
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "27",
                "Q1",
                account_id,
                contact_id,
                "Smoke Test Play",
                "Smoke Test Play",
                "2026-05-02",
                "09:00",
                "Call",
                "Smoke historical due edit source",
                "No Response",
                "Historical due remains unchanged",
                "2026-05-05",
                "10:00",
                "Not Started",
                "Smoke Test Admin",
            ),
        ).lastrowid
        connection.commit()
        connection.close()
        response = client.post(
            f"/outreach/{historical_due_outreach_id}/edit",
            data={
                "csrf_token": csrf_from_session(client),
                "submit_action": "save",
                "fy": "27",
                "quarter": "Q1",
                "account_id": str(account_id),
                "sales_play": "Smoke Test Play",
                "contact_ids": [str(contact_id)],
                "task_status": "Not Started",
                "assigned_to": "Smoke Test Admin",
                "activity_type": "Call",
                "activity_date": "2026-04-01",
                "activity_time": "08:00",
                "next_action_date": "2026-05-05",
                "next_action_time": "10:00",
                "subject": "Smoke historical due edit source",
                "outcome": "No Response",
                "scheduled_meeting_at": "",
                "next_action": "Backdated activity start while retaining old due date",
            },
            follow_redirects=False,
        )
        assert_ok(
            response.status_code in (302, 303),
            "editing Activity Start was blocked by an unchanged historical due date",
        )

        connection = sqlite3.connect(db_path)
        connection.execute("ALTER TABLE outreach_recipients DROP COLUMN sort_order")
        connection.commit()
        connection.close()
        response = client.get(f"/outreach/{multi_outreach['id']}/edit")
        edit_html = response.get_data(as_text=True)
        assert_ok(
            response.status_code == 200
            and "class=\"checkbox-select\"" in edit_html
            and "name=\"contact_ids\"" in edit_html,
            "edit Outreach multi-contact checkbox field missing",
        )

        response = client.post(
            f"/outreach/{multi_outreach['id']}/edit",
            data={
                "csrf_token": csrf_from_session(client),
                "submit_action": "save",
                "fy": "27",
                "quarter": "Q1",
                "account_id": str(account_id),
                "sales_play": "Smoke Test Play",
                "contact_ids": [str(second_contact_id)],
                "task_status": "In Progress",
                "assigned_to": "Smoke Test Admin",
                "activity_type": "Call",
                "activity_date": "2026-05-07",
                "activity_time": "09:00",
                "next_action_date": "2026-08-10",
                "next_action_time": "10:00",
                "subject": "Smoke multi-contact outreach",
                "outcome": "Exec Meeting Booked",
                "scheduled_meeting_at": f"{future_meeting_date}T10:30",
                "next_action": "",
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "multi-contact outreach edit failed")

        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        edited_outreach = connection.execute(
            "SELECT * FROM outreach WHERE id = ?",
            (multi_outreach["id"],),
        ).fetchone()
        edited_recipient_count = connection.execute(
            "SELECT COUNT(*) FROM outreach_recipients WHERE outreach_id = ?",
            (multi_outreach["id"],),
        ).fetchone()[0]
        connection.close()
        assert_ok(str(edited_outreach["contact_id"]) == str(second_contact_id), "edited primary contact was not updated")
        assert_ok(edited_outreach["outcome"] == "Exec Meeting Booked", "Exec Meeting Booked outcome was not saved")
        assert_ok(edited_recipient_count == 1, "edited outreach recipients were not replaced")

        response = client.post(
            "/outreach/add",
            data={
                "csrf_token": csrf_from_session(client),
                "fy": "27",
                "quarter": "Q1",
                "account_id": str(account_id),
                "sales_play": "Smoke Test Play",
                "contact_ids": [str(contact_id)],
                "task_status": "Not Started",
                "assigned_to": "Smoke Test Admin",
                "activity_type": "Email",
                "activity_date": "2026-05-14",
                "activity_time": "08:30",
                "next_action_date": "2026-08-17",
                "next_action_time": "09:45",
                "subject": "Smoke complete follow-on source",
                "outcome": "No Response",
                "next_action": "Prepare completion update",
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "complete follow-on source outreach add failed")
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        follow_source = connection.execute(
            "SELECT id FROM outreach WHERE subject = ?",
            ("Smoke complete follow-on source",),
        ).fetchone()
        connection.execute("ALTER TABLE outreach DROP COLUMN scheduled_meeting_time")
        connection.commit()
        connection.close()
        assert_ok(follow_source is not None, "complete follow-on source outreach was not saved")

        response = client.post(
            f"/outreach/{follow_source['id']}/edit",
            data={
                "csrf_token": csrf_from_session(client),
                "submit_action": "complete_and_follow",
                "fy": "27",
                "quarter": "Q1",
                "account_id": str(account_id),
                "sales_play": "Smoke Test Play",
                "contact_ids": [str(contact_id)],
                "task_status": "In Progress",
                "assigned_to": "Smoke Test Admin",
                "activity_type": "Email",
                "activity_date": "2026-05-14",
                "activity_time": "08:30",
                "next_action_date": "2026-08-17",
                "next_action_time": "09:45",
                "subject": "Smoke complete follow-on source",
                "outcome": "No Response",
                "next_action": "Completed and creating follow on",
            },
            follow_redirects=False,
        )
        assert_ok(
            response.status_code in (302, 303) and f"/outreach/add?prefill_from={follow_source['id']}" in response.headers.get("Location", ""),
            "complete and create follow-on did not redirect to the prefilled new Outreach form",
        )
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        completed_follow_source = connection.execute(
            "SELECT task_status, scheduled_meeting_time FROM outreach WHERE id = ?",
            (follow_source["id"],),
        ).fetchone()
        connection.close()
        assert_ok(completed_follow_source["task_status"] == "Completed", "complete and follow-on source was not completed")

        response = client.get("/outreach")
        outreach_html = response.get_data(as_text=True)
        assert_ok("bulk_next_action_date" in outreach_html, "bulk Outreach due-date control missing")
        assert_ok("next_action_date_" in outreach_html, "inline Outreach due-date control missing")
        assert_ok('data-select-all="bulk-outreach-form"' in outreach_html, "bulk Outreach select-all control missing")
        assert_ok("outreach-auto-reschedule-form" in outreach_html, "row Outreach auto-reschedule control missing")
        assert_ok("Reschedule Selected" in outreach_html, "bulk Outreach reschedule action missing")
        assert_ok("outreach-bulk-reschedule-button" in outreach_html, "bulk reschedule action is not visually distinguished")

        connection = sqlite3.connect(db_path)
        connection.execute("DROP TABLE timeline_entries")
        connection.commit()
        connection.close()
        inline_manual_past_due_date = (today - timedelta(days=7)).isoformat()
        response = client.post(
            f"/outreach/{multi_outreach['id']}/due-date",
            data={
                "csrf_token": csrf_from_session(client),
                "return_to": "/outreach",
                "next_action_date": inline_manual_past_due_date,
                "next_action_time": "12:15",
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "inline Outreach due-date update failed")

        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        inline_due = connection.execute(
            "SELECT next_action_date, next_action_time FROM outreach WHERE id = ?",
            (multi_outreach["id"],),
        ).fetchone()
        connection.close()
        assert_ok(inline_due["next_action_date"] == inline_manual_past_due_date, "inline retrospective due date was not saved")
        assert_ok(inline_due["next_action_time"] == "12:15", "inline due time was not saved")

        blocked_contact_dates = []
        candidate_day = date.today()
        while len(blocked_contact_dates) < 3:
            if candidate_day.weekday() < 5:
                blocked_contact_dates.append(candidate_day.isoformat())
            candidate_day += timedelta(days=1)
        connection = sqlite3.connect(db_path)
        for index, blocked_date in enumerate(blocked_contact_dates, start=1):
            connection.execute(
                """
                INSERT INTO outreach (
                    fy, quarter, account_id, contact_id, campaign, sales_play,
                    activity_date, activity_time, activity_type, subject, outcome,
                    next_action, next_action_date, next_action_time, task_status, assigned_to
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "27",
                    "Q1",
                    account_id,
                    second_contact_id,
                    "Smoke Test Play",
                    "Smoke Test Play",
                    blocked_date,
                    "09:00",
                    "Email",
                    f"Smoke contact date blocker {index}",
                    "No Response",
                    "Blocking same-contact date",
                    blocked_date,
                    "09:00",
                    "Not Started",
                    "Smoke Test Admin",
                ),
            )
        connection.commit()
        connection.close()

        response = client.post(
            f"/outreach/{multi_outreach['id']}/auto-reschedule",
            data={
                "csrf_token": csrf_from_session(client),
                "return_to": "/outreach",
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "single Outreach auto-reschedule failed")
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        auto_due = connection.execute(
            "SELECT next_action_date, next_action_time FROM outreach WHERE id = ?",
            (multi_outreach["id"],),
        ).fetchone()
        connection.close()
        assert_ok(auto_due["next_action_date"] and auto_due["next_action_time"], "single auto-reschedule did not set a due slot")
        assert_ok(date.fromisoformat(auto_due["next_action_date"]).weekday() < 5, "single auto-reschedule selected a weekend")
        assert_ok(auto_due["next_action_date"] not in blocked_contact_dates, "single auto-reschedule selected a date already used by the same contact")

        first_original_due = (today + timedelta(days=31)).isoformat()
        second_original_due = (today + timedelta(days=30)).isoformat()
        connection = sqlite3.connect(db_path)
        connection.execute(
            "UPDATE outreach SET next_action_date = ?, next_action_time = '11:00' WHERE id = ?",
            (first_original_due, outreach_id),
        )
        connection.execute(
            "UPDATE outreach SET next_action_date = ?, next_action_time = '10:00' WHERE id = ?",
            (second_original_due, multi_outreach["id"]),
        )
        connection.commit()
        connection.close()

        response = client.post(
            "/outreach/bulk-action",
            data={
                "csrf_token": csrf_from_session(client),
                "return_to": "/outreach",
                "bulk_action": "auto_reschedule",
                "selected_ids": [str(outreach_id), str(multi_outreach["id"])],
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "bulk Outreach auto-reschedule failed")
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        auto_bulk_rows = connection.execute(
            "SELECT id, next_action_date, next_action_time FROM outreach WHERE id IN (?, ?)",
            (outreach_id, multi_outreach["id"]),
        ).fetchall()
        other_schedule_rows = connection.execute(
            """
            SELECT id, activity_date, activity_time, next_action_date, next_action_time
            FROM outreach
            WHERE COALESCE(task_status, '') NOT IN ('Deleted', 'Cancelled')
              AND id NOT IN (?, ?)
            """,
            (outreach_id, multi_outreach["id"]),
        ).fetchall()
        connection.close()
        auto_bulk_slots = {(row["next_action_date"], row["next_action_time"]) for row in auto_bulk_rows}
        assert_ok(len(auto_bulk_rows) == 2 and len(auto_bulk_slots) == 2, "bulk auto-reschedule created clashing due slots")
        assert_ok(all(date.fromisoformat(row["next_action_date"]).weekday() < 5 for row in auto_bulk_rows), "bulk auto-reschedule selected a weekend")
        bulk_by_id = {row["id"]: row for row in auto_bulk_rows}
        earlier_slot = datetime.fromisoformat(
            f"{bulk_by_id[multi_outreach['id']]['next_action_date']}T{bulk_by_id[multi_outreach['id']]['next_action_time']}"
        )
        later_slot = datetime.fromisoformat(
            f"{bulk_by_id[outreach_id]['next_action_date']}T{bulk_by_id[outreach_id]['next_action_time']}"
        )
        assert_ok(earlier_slot < later_slot, "bulk auto-reschedule did not preserve the tasks' original chronological order")
        assert_ok((later_slot - earlier_slot).total_seconds() >= 15 * 60, "bulk auto-reschedule did not keep tasks 15 minutes apart")
        other_slots = []
        for row in other_schedule_rows:
            if row["activity_date"]:
                other_slots.append(datetime.fromisoformat(f"{row['activity_date']}T{(row['activity_time'] or '09:00')[:5]}"))
            if row["next_action_date"]:
                other_slots.append(datetime.fromisoformat(f"{row['next_action_date']}T{(row['next_action_time'] or '09:00')[:5]}"))
        for selected_slot in (earlier_slot, later_slot):
            assert_ok(
                all(
                    selected_slot.date() != occupied.date()
                    or abs((selected_slot - occupied).total_seconds()) >= 15 * 60
                    for occupied in other_slots
                ),
                "bulk auto-reschedule clashed with an existing Outreach schedule slot",
            )

        connection = sqlite3.connect(db_path)
        connection.execute("ALTER TABLE outreach DROP COLUMN scheduled_meeting_time")
        connection.commit()
        connection.close()
        response = client.post(
            "/outreach/add",
            data={
                "csrf_token": csrf_from_session(client),
                "fy": "27",
                "quarter": "Q1",
                "account_id": str(account_id),
                "sales_play": "Smoke Test Play",
                "contact_ids": [str(contact_id)],
                "task_status": "Not Started",
                "assigned_to": "Smoke Test Admin",
                "activity_type": "Email",
                "activity_date": "2026-05-11",
                "activity_time": "08:45",
                "next_action_date": "2026-08-19",
                "next_action_time": "09:15",
                "subject": "Smoke schema recovery outreach",
                "outcome": "No Response",
                "next_action": "Try a different route",
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "schema recovery outreach add failed")
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        recovered_outreach = connection.execute(
            "SELECT scheduled_meeting_time FROM outreach WHERE subject = ?",
            ("Smoke schema recovery outreach",),
        ).fetchone()
        connection.close()
        assert_ok(recovered_outreach is not None, "schema recovery outreach was not saved")

        bulk_manual_past_due_date = (today - timedelta(days=3)).isoformat()
        response = client.post(
            "/outreach/bulk-action",
            data={
                "csrf_token": csrf_from_session(client),
                "return_to": "/outreach",
                "bulk_action": "update_due",
                "selected_ids": [str(outreach_id), str(multi_outreach["id"])],
                "bulk_next_action_date": bulk_manual_past_due_date,
                "bulk_next_action_time": "14:45",
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "bulk Outreach due-date update failed")

        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        bulk_due_rows = connection.execute(
            "SELECT next_action_date, next_action_time FROM outreach WHERE id IN (?, ?) ORDER BY id",
            (outreach_id, multi_outreach["id"]),
        ).fetchall()
        connection.close()
        assert_ok(
            all(row["next_action_date"] == bulk_manual_past_due_date and row["next_action_time"] == "14:45" for row in bulk_due_rows),
            "bulk retrospective due date was not saved for selected outreach tasks",
        )

        task_manual_past_due_date = (today - timedelta(days=5)).isoformat()
        response = client.post(
            f"/tasks/{outreach_id}/update",
            data={
                "csrf_token": csrf_from_session(client),
                "return_to": "/tasks",
                "outcome": "No Response",
                "task_status": "In Progress",
                "next_action": "Manual retrospective task schedule smoke coverage.",
                "next_action_date": task_manual_past_due_date,
                "next_action_time": "09:30",
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "Tasks page manually entered retrospective due date was blocked")
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        task_due_row = connection.execute(
            "SELECT next_action_date, next_action_time FROM outreach WHERE id = ?",
            (outreach_id,),
        ).fetchone()
        connection.close()
        assert_ok(
            task_due_row
            and task_due_row["next_action_date"] == task_manual_past_due_date
            and task_due_row["next_action_time"] == "09:30",
            "Tasks page manually entered retrospective due date was not saved",
        )

        for path in (
            "/reports/accounts/export",
            "/reports/accounts/export?format=xlsx",
            "/reports/contacts/export",
            "/reports/contacts/export?format=xlsx",
            "/reports/outreach/export",
            "/reports/outreach/export?format=xlsx",
            "/reports/partners/export",
            "/reports/partners/export?format=xlsx",
            "/reports/tasks/export",
            "/reports/full-export.xlsx",
            "/reports/pg-progress/export.xls",
            "/reports/pg-progress/export.pdf",
            "/outreach/export",
        ):
            response = client.get(path)
            assert_ok(response.status_code == 200, f"{path} returned {response.status_code}")
            assert_ok(response.headers.get("Content-Disposition"), f"{path} did not download")

        response = client.get("/reports")
        assert_ok(response.status_code == 200, f"Reports page returned {response.status_code}")
        assert_ok(b'id="pgBibleExportDialog"' in response.data, "PG Bible period selection dialog was not rendered")
        assert_ok(b'name="fy"' in response.data and b'name="quarter"' in response.data, "PG Bible FY and Quarter selectors were not rendered")

        response = client.get("/reports/pg-bible/export", follow_redirects=False)
        assert_ok(response.status_code in (302, 303), "PG Bible export without FY and Quarter was not redirected")

        response = client.get("/reports/pg-bible/export?fy=FY27&quarter=Q1")
        assert_ok(response.status_code == 200, f"PG Bible export returned {response.status_code}")
        assert_ok(response.headers.get("Content-Disposition"), "PG Bible did not download")
        from excel_exporter import PGBibleExporter
        from openpyxl import load_workbook
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
        report = pipeflow_app.build_pg_bible_report_from_db(connection, ("FY27",), ("Q1",))
        excluded_period_report = pipeflow_app.build_pg_bible_report_from_db(connection, ("FY27",), ("Q4",))
        connection.close()
        assert_ok(not excluded_period_report.plan_items, "PG Bible period filters included accounts from an unselected quarter")
        assert_ok(not excluded_period_report.action_items, "PG Bible period filters included actions from an unselected quarter")
        pg_bible_output = PGBibleExporter(
            Path(pipeflow_app.__file__).resolve().parent / "pg_bible_templates" / "PGBible_Template_May2026.xlsx",
            Path(tmp) / "pg_bible_exports",
        ).export(report)
        pg_workbook = load_workbook(pg_bible_output, data_only=False)
        pg_sheet = pg_workbook.active
        assert_ok(pg_sheet["B5"].value is not None, "PG Bible Pipeline Added was not written to the merged input anchor")
        assert_ok(pg_sheet["L5"].value == "=(F3+B5)-L3", "PG Bible pipeline gap formula was not adapted for the unchanged merged template")
        assert_ok(pg_sheet["L11"].value == "Smoke Test Account - BMC", "PG Bible PG PLAN customer did not include business organisation")
        assert_ok(pg_sheet["D11"].value == "Smoke Test Play", "PG Bible PG PLAN Sales Play did not use configured account Sales Play")
        assert_ok(str(pg_sheet["B11"].value) == "1", "PG Bible PG PLAN NBM target did not map Account PG Bible Order")
        pg_action_row = next(
            row
            for row in range(33, 80)
            if (pg_sheet[f"C{row}"].value or "").startswith("Smoke Test Contact")
        )
        assert_ok("Smoke Test Account - BMC" in (pg_sheet[f"C{pg_action_row}"].value or ""), "PG Bible PG ACTIONS contact label missing account context")
        assert_ok(str(pg_sheet[f"B{pg_action_row}"].value) == "1", "PG Bible PG ACTIONS target did not map Account PG Bible Order")
        assert_ok(pg_sheet[f"F{pg_action_row}"].value == "Yes", "PG Bible Discovery Meeting did not map from a booked meeting with a scheduled date")
        assert_ok("08-05-2026" in (pg_sheet[f"J{pg_action_row}"].value or ""), "PG Bible NBM booked column did not include scheduled meeting date")
        assert_ok(pipeflow_app.money_value(pg_sheet[f"T{pg_action_row}"].value) > 0, "PG Bible VO Value did not map for qualified scheduled meeting progress")

        from models import ActionItem, OwnerReport, PlanItem
        expanded_report = OwnerReport(
            profile=report.profile,
            goals=report.goals,
            calc_payload=report.calc_payload,
            plan_items=[
                PlanItem(
                    pg_bible_order=index,
                    nbm_target=str(index),
                    customer=f"Capacity Account {index}",
                )
                for index in range(1, 22)
            ],
            action_items=[
                ActionItem(
                    related_nbm_target=str((index % 21) + 1),
                    discovery_target_name_title=f"Capacity Contact {index + 1}",
                )
                for index in range(50)
            ],
        )
        expanded_output = PGBibleExporter(
            Path(pipeflow_app.__file__).resolve().parent / "pg_bible_templates" / "PGBible_Template_May2026.xlsx",
            Path(tmp) / "pg_bible_capacity_exports",
        ).export(expanded_report)
        expanded_sheet = load_workbook(expanded_output, data_only=False).active
        assert_ok(str(expanded_sheet["B31"].value) == "21", "PG Bible did not extend PG PLAN NBM target mapping beyond row 29")
        assert_ok(expanded_sheet["B33"].value == "PG ACTIONS", "Extended PG PLAN rows did not preserve the PG ACTIONS section")
        assert_ok(str(expanded_sheet["B35"].value) == "1", "Extended PG ACTIONS target mapping did not start in the shifted action row")
        assert_ok(str(expanded_sheet["B84"].value) == "8", "PG Bible did not extend NBM target mapping through every action row")

        app_source = Path("app.py").read_text(encoding="utf-8")
        account_templates = (
            Path("templates/add_account.html").read_text(encoding="utf-8")
            + Path("templates/edit_account.html").read_text(encoding="utf-8")
        )
        assert_ok('request.form.get("nbm_target")' not in app_source, "Removed Account NBM Target is still accepted by save handlers")
        assert_ok('name="nbm_target"' not in account_templates, "Removed Account NBM Target is still displayed on Account forms")

        response = client.post(
            "/outreach/bulk-action",
            data={
                "csrf_token": csrf_from_session(client),
                "return_to": "/outreach",
                "bulk_action": "delete",
                "selected_ids": [str(outreach_id), str(multi_outreach["id"])],
            },
            follow_redirects=False,
        )
        assert_ok(response.status_code in (302, 303), "bulk Outreach delete failed")

        connection = sqlite3.connect(db_path)
        remaining_outreach = connection.execute(
            "SELECT COUNT(*) FROM outreach WHERE id IN (?, ?)",
            (outreach_id, multi_outreach["id"]),
        ).fetchone()[0]
        remaining_recipients = connection.execute(
            "SELECT COUNT(*) FROM outreach_recipients WHERE outreach_id IN (?, ?)",
            (outreach_id, multi_outreach["id"]),
        ).fetchone()[0]
        connection.close()
        assert_ok(remaining_outreach == 0, "bulk Outreach delete did not remove selected tasks")
        assert_ok(remaining_recipients == 0, "bulk Outreach delete did not remove selected recipients")

        assert_ok(
            "SELECT MAX(id) AS id FROM outreach" not in app_source,
            "Outreach inserts must not infer the saved task with SELECT MAX(id)",
        )

    print("PipeFlow smoke test passed.")


if __name__ == "__main__":
    main()
