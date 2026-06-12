from __future__ import annotations

import copy
import logging
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

for vendor_base in (
    Path(getattr(sys, "_MEIPASS", Path(__file__).parent)),
    Path(__file__).parent,
    Path(sys.executable).resolve().parent.parent / "Resources",
):
    local_vendor = vendor_base / "vendor"
    if local_vendor.exists():
        vendor_path = str(local_vendor)
        if vendor_path not in sys.path:
            sys.path.insert(0, vendor_path)

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models import GoalsSummary, OwnerReport, PGBibleExportError


LOGGER = logging.getLogger(__name__)
SECTION_LABELS = ["PG GOALS", "PG PLAN", "PG ACTIONS", "PG RESULTS"]
REQUIRED_SECTION_LABELS = ["PG GOALS", "PG PLAN", "PG ACTIONS"]
INVALID_SHEET_CHARS = r"[]:*?/\\"
MONTH_ORDER = {
    "april": 1,
    "may": 2,
    "june": 3,
    "july": 4,
    "august": 5,
    "september": 6,
    "october": 7,
    "november": 8,
    "december": 9,
    "january": 10,
    "february": 11,
    "march": 12,
}
QUARTER_MARKERS = {"Q1 Results", "Q2 Results", "Q3 Results", "Q4 Results"}
PLAN_ENTRY_ROWS = range(11, 30)
MONTH_PLAN_ROWS = {
    "april": 9,
    "may": 10,
    "june": 11,
    "july": 12,
    "august": 13,
    "september": 14,
    "october": 15,
    "november": 16,
    "december": 17,
    "january": 18,
    "february": 19,
    "march": 20,
}
ACTION_ENTRY_ROWS = range(33, 80)
NBM_COLOUR_PALETTE = {
    0: ("D90000", "FFFFFF"),
    1: ("F00000", "FFFFFF"),
    2: ("FFC000", "FFFFFF"),
    3: ("FFF200", "111111"),
    4: ("92D050", "FFFFFF"),
    5: ("00B050", "FFFFFF"),
    6: ("00B0F0", "FFFFFF"),
    7: ("0070C0", "FFFFFF"),
    8: ("002060", "FFFFFF"),
    9: ("000000", "FFFFFF"),
    10: ("7F7F7F", "FFFFFF"),
    11: ("595959", "FFFFFF"),
}


@dataclass
class Section:
    label: str
    row: int
    column: int


@dataclass
class TableRegion:
    name: str
    header_rows: set[int]
    columns: dict[str, int]
    start_row: int
    end_row: int


@dataclass
class WeeklyKey:
    column: int
    header: str
    data_type: str


def norm(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def sanitize_excel_name(name: str) -> str:
    cleaned = "".join(ch for ch in (name or "PipeFlow") if ch not in INVALID_SHEET_CHARS).strip()
    return (cleaned or "PipeFlow")[:31]


def sanitize_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value or "PipeFlow").strip("_")
    return cleaned or "PipeFlow"


def decimal_value(value: Any) -> Decimal:
    if value in (None, ""):
        raise PGBibleExportError("CALC_INPUT_MISSING", "A required calculation input is missing.", ["empty value"])
    return Decimal(str(value).replace("$", "").replace(",", ""))


def compute_starting_pipeline(reporting_date: date, pipeflow_payload: dict[str, Any]) -> Decimal:
    if "starting_pipeline" not in pipeflow_payload:
        raise PGBibleExportError(
            "CALC_INPUT_MISSING",
            "Starting pipeline cannot be calculated because required inputs are missing.",
            ["starting_pipeline"],
        )
    return decimal_value(pipeflow_payload["starting_pipeline"])


def compute_pipeline_added(fy_start: date, reporting_date: date, pipeflow_payload: dict[str, Any]) -> Decimal:
    if "pipeline_added" not in pipeflow_payload:
        raise PGBibleExportError(
            "CALC_INPUT_MISSING",
            "Pipeline added cannot be calculated because required inputs are missing.",
            ["pipeline_added"],
        )
    return decimal_value(pipeflow_payload["pipeline_added"])


def compute_pipeline_target(fy_window: str, pipeflow_payload: dict[str, Any]) -> Decimal:
    if "pipeline_target" not in pipeflow_payload:
        raise PGBibleExportError(
            "CALC_INPUT_MISSING",
            "Pipeline target cannot be calculated because required inputs are missing.",
            ["pipeline_target"],
        )
    return decimal_value(pipeflow_payload["pipeline_target"])


def compute_pipeline_gap(target: Decimal, starting: Decimal, added: Decimal) -> Decimal:
    return (starting + added) - target


class PGBibleExporter:
    def __init__(self, template_path: str | Path, output_dir: str | Path = "."):
        self.template_path = Path(template_path)
        self.output_dir = Path(output_dir)
        self.sections: dict[str, Section] = {}
        self.header_cache: dict[str, TableRegion] = {}
        self.weekly_key: WeeklyKey | None = None

    def export(self, report: OwnerReport, reporting_date: date | None = None) -> Path:
        reporting_date = reporting_date or date.today()
        wb = load_workbook(self.template_path, data_only=False)

        if "Real example" in wb.sheetnames:
            ws = wb["Real example"]
        elif wb.worksheets:
            ws = wb.worksheets[0]
        else:
            raise PGBibleExportError(
                "TEMPLATE_SHEET_MISSING",
                "The template workbook is missing a writable PG Bible worksheet.",
                ["Real example or first worksheet"],
            )
        for other in list(wb.worksheets):
            if other is not ws:
                wb.remove(other)

        final_sheet_name = sanitize_excel_name(report.profile.profile_name)
        ws.title = final_sheet_name
        print(f"profile name used: {report.profile.profile_name}")
        print(f"sheet name final: {final_sheet_name}")

        self._discover_template(ws)
        baseline = self._structural_snapshot(ws)
        report.goals = report.goals or self._compute_goals(report, reporting_date)

        self._clear_template_inputs(ws)
        if "PG RESULTS" in self.header_cache:
            self._clear_weekly_rows(ws, self.header_cache["PG RESULTS"])

        self._write_goals(ws, report.goals)
        plan_count = self._write_plan(ws, report)
        monthly_count = self._write_monthly_plan(ws, report)
        action_count = self._write_actions(ws, report)
        weekly_count = self._write_weekly_results(ws, report) if "PG RESULTS" in self.header_cache else 0

        print(f"plan rows written: {plan_count}")
        print(f"monthly plan rows written: {monthly_count}")
        print(f"action rows written: {action_count}")
        print(f"weekly rows written: {weekly_count}")

        output_path = self.output_dir / f"PGBible_{sanitize_filename(report.profile.username)}.xlsx"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self._self_check(ws, baseline)
        wb.save(output_path)
        saved_wb = load_workbook(output_path, data_only=False)
        self._self_check(saved_wb[final_sheet_name], baseline)
        return output_path

    def _compute_goals(self, report: OwnerReport, reporting_date: date) -> GoalsSummary:
        fy_start = date(reporting_date.year if reporting_date.month >= 4 else reporting_date.year - 1, 4, 1)
        fy_window = f"{fy_start.isoformat()} to {date(fy_start.year + 1, 3, 31).isoformat()}"
        starting = compute_starting_pipeline(reporting_date, report.calc_payload)
        added = compute_pipeline_added(fy_start, reporting_date, report.calc_payload)
        target = compute_pipeline_target(fy_window, report.calc_payload)
        gap = compute_pipeline_gap(target, starting, added)
        print(f"calculation starting_pipeline input: {starting}")
        print(f"calculation pipeline_added input: {added}")
        print(f"calculation pipeline_target input: {target}")
        print(f"calculation pipeline_gap output: {gap}")
        return GoalsSummary(starting_pipeline=starting, pipeline_added=added, pipeline_target=target, pipeline_gap=gap)

    def _discover_template(self, ws) -> None:
        self._validate_sections(ws)
        self._validate_may_2026_layout(ws)
        if "PG RESULTS" in self.sections:
            try:
                self.header_cache["PG RESULTS"] = self._discover_results(ws)
                self.weekly_key = self._discover_weekly_key(ws, self.header_cache["PG RESULTS"])
                print(f"Weekly key resolved as: {get_column_letter(self.weekly_key.column)}, data type: {self.weekly_key.data_type}")
            except PGBibleExportError as exc:
                LOGGER.warning("PG RESULTS section present but not writable: %s %s", exc.error_code, exc.details)
                self.header_cache.pop("PG RESULTS", None)
                self.weekly_key = None

    def _validate_sections(self, ws) -> None:
        for label in SECTION_LABELS:
            matches = self._find_exact(ws, label)
            if not matches:
                if label in REQUIRED_SECTION_LABELS:
                    LOGGER.warning("PG Bible template is missing expected section %s; that block will be skipped.", label)
            elif len(matches) > 1:
                LOGGER.warning("PG Bible template section %s is ambiguous; using first occurrence.", label)
                cell = matches[0]
                self.sections[label] = Section(label, cell.row, cell.column)
            else:
                cell = matches[0]
                self.sections[label] = Section(label, cell.row, cell.column)

    def _validate_may_2026_layout(self, ws) -> None:
        expected = {
            "B2": "PG GOALS",
            "B8": "PG PLAN",
            "B31": "PG ACTIONS",
            "B9": "NBM Target",
            "D9": "PG Sales Play or Initative",
            "L9": "Customer",
            "M9": "Estimated Value",
            "O8": "Month",
            "P8": "PG Marketing Event Being Supported",
            "S8": "Notes & General PG Actions",
            "B32": "Related NBM Target",
            "C32": "Account / Contact",
            "F32": "Completed Discovery Meeting",
            "G32": "Next Action / Notes",
            "J32": "NBM Booked / Date",
            "M32": "Why Buy Document",
            "N32": "Exec First",
            "O32": "Preparation With Manager",
            "P32": "Completed NBM",
            "Q32": "NBM Next Action",
            "T32": "VO Value",
        }
        for coordinate, expected_text in expected.items():
            actual = self._merged_value(ws, ws[coordinate].row, ws[coordinate].column)
            if expected_text.casefold() not in str(actual or "").casefold():
                LOGGER.warning(
                    "PG Bible template cell %s was expected to contain %r but contains %r.",
                    coordinate,
                    expected_text,
                    actual,
                )

    def _discover_plan(self, ws) -> TableRegion:
        return self._discover_header_block(
            ws,
            "PG PLAN",
            "PG ACTIONS",
            {
                "month": ["Month"],
                "marketing_event": ["PG Marketing Event Being Supported"],
                "notes": ["Notes & General PG Actions"],
                "nbm_target": ["NBM Target"],
                "sales_play": ["PG Sales Play or Initative", "PG Sales Play or Initiative"],
                "customer": ["Customer"],
                "estimated_value": ["Estimated Value"],
            },
        )

    def _discover_actions(self, ws) -> TableRegion:
        return self._discover_header_block(
            ws,
            "PG ACTIONS",
            "PG RESULTS" if "PG RESULTS" in self.sections else None,
            {
                "related_nbm_target": ["Related NBM Target"],
                "discovery_target_name_title": ["Targeted or Booked Discovery Meeting (Name & Poistion)", "Targeted or Booked Discovery Meeting (Name & Position)", "Account / Contact"],
                "discovery_completed": ["Completed Discovery Meeting Yes / No"],
                "discovery_next_action": ["Discovery Meeting Next Action / Notes", "Next Action / Notes"],
                "nbm_booked": ["NBM Booked / Date (Name & Position)"],
                "why_buy": ["Why Buy Document Yes / No"],
                "exec_first": ["Exec First Yes / No"],
                "prep_with_manager": ["Preparation With Manager"],
                "nbm_completed": ["Completed NBM Yes / No"],
                "nbm_next_action": ["NBM Next Action / \\Notes", "NBM Next Action / Notes"],
                "vo_value": ["VO Value"],
            },
        )

    def _discover_results(self, ws) -> TableRegion:
        return self._discover_header_block(
            ws,
            "PG RESULTS",
            None,
            {
                "week_number": ["#"],
                "week_commencing": ["WC"],
                "vitos_sent": ["Sent"],
                "vitos_chased": ["Chased"],
                "discovery_booked": ["Booked"],
                "discovery_completed": ["Completed"],
                "nbms_booked": ["Booked"],
                "nbms_exec_firsts": ["Exec Firsts"],
                "nbms_completed": ["Completed"],
                "pipeline_generated_vo_count": ["#VO"],
                "pipeline_generated_value": ["$m"],
            },
            allow_duplicate_labels=True,
        )

    def _discover_header_block(self, ws, section_name: str, next_section_name: str | None, expected: dict[str, list[str]], allow_duplicate_labels: bool = False) -> TableRegion:
        start = self.sections[section_name].row
        end = self.sections[next_section_name].row - 1 if next_section_name else ws.max_row
        search_end = min(end, start + 5)
        columns: dict[str, int] = {}
        header_rows: set[int] = set()
        used_columns: set[int] = set()

        for key, labels in expected.items():
            matches = []
            for row in range(start, search_end + 1):
                for col in range(1, ws.max_column + 1):
                    value = self._merged_value(ws, row, col)
                    if any(norm(value) == norm(label) for label in labels):
                        matches.append((row, col))
            if not matches:
                raise PGBibleExportError("TABLE_HEADER_NOT_FOUND", "A required table header could not be found.", [f"{section_name}: {labels[0]}"])
            chosen = None
            for match in matches:
                if allow_duplicate_labels and match[1] in used_columns:
                    continue
                chosen = match
                break
            if chosen is None:
                chosen = matches[0]
            header_rows.add(chosen[0])
            columns[key] = chosen[1]
            used_columns.add(chosen[1])

        start_row = max(header_rows) + 1
        for merged in ws.merged_cells.ranges:
            if any(merged.min_row <= header_row <= merged.max_row for header_row in header_rows):
                start_row = max(start_row, merged.max_row + 1)

        return TableRegion(section_name, header_rows, columns, start_row, end)

    def _discover_weekly_key(self, ws, region: TableRegion) -> WeeklyKey:
        rows = self._weekly_data_rows(ws, region)
        candidates: list[WeeklyKey] = []
        for col in region.columns.values():
            values = [ws.cell(row, col).value for row in rows if ws.cell(row, col).value not in (None, "")]
            if not values or len(values) != len(set(str(v) for v in values)):
                continue
            if all(isinstance(value, (datetime, date)) for value in values):
                candidates.append(WeeklyKey(col, self._header_for_column(ws, region, col), "date"))
            elif all(isinstance(value, int) and 1 <= value <= 53 for value in values):
                candidates.append(WeeklyKey(col, self._header_for_column(ws, region, col), "int"))

        date_candidates = [candidate for candidate in candidates if candidate.data_type == "date"]
        int_candidates = [candidate for candidate in candidates if candidate.data_type == "int"]
        if len(date_candidates) == 1:
            return date_candidates[0]
        if len(date_candidates) > 1:
            raise PGBibleExportError("WEEK_KEY_AMBIGUOUS", "Multiple weekly date key columns were found.", [c.header for c in date_candidates])
        if len(int_candidates) == 1:
            return int_candidates[0]
        raise PGBibleExportError("WEEK_KEY_AMBIGUOUS", "The weekly key column could not be resolved.", [c.header for c in candidates])

    def _write_goals(self, ws, goals: GoalsSummary) -> None:
        if "PG GOALS" not in self.sections:
            return
        self._write_coordinate(ws, "F3", goals.starting_pipeline)
        self._write_coordinate(ws, "L3", goals.pipeline_target)
        # The May 2026 template has no visible "Pipeline Added" label, but L5
        # calculates the gap with =(F3+F5)-L3, making F5 the intended input.
        # F5 is inside a merged area, so _write_coordinate stores the value on
        # the merged range anchor while preserving the workbook structure.
        self._write_coordinate(ws, "F5", goals.pipeline_added)

    def _write_plan(self, ws, report: OwnerReport) -> int:
        if "PG PLAN" not in self.sections:
            return 0
        rows = sorted(
            report.plan_items,
            key=lambda item: (
                item.pg_bible_order if item.pg_bible_order is not None else 999999,
                item.customer.casefold(),
            ),
        )
        writable_rows = list(PLAN_ENTRY_ROWS)
        if len(rows) > len(writable_rows):
            LOGGER.warning("PG Bible plan has %s accounts but only %s rows; extra accounts were not exported.", len(rows), len(writable_rows))
        for row, item in zip(writable_rows, rows):
            self._write_nbm_target(ws, f"B{row}", item.nbm_target)
            self._write_coordinate(ws, f"D{row}", item.sales_play)
            self._write_coordinate(ws, f"L{row}", item.customer)
            self._write_coordinate(ws, f"M{row}", item.estimated_value)
        return min(len(rows), len(writable_rows))

    def _write_monthly_plan(self, ws, report: OwnerReport) -> int:
        if "PG PLAN" not in self.sections:
            return 0
        items = {norm(item.month): item for item in report.monthly_plan_items if norm(item.month)}
        written = 0
        for month_key, row in MONTH_PLAN_ROWS.items():
            item = items.get(month_key)
            if not item:
                continue
            self._write_coordinate(ws, f"P{row}", item.marketing_event)
            self._write_coordinate(ws, f"S{row}", item.notes)
            written += 1
        return written

    def _write_actions(self, ws, report: OwnerReport) -> int:
        if "PG ACTIONS" not in self.sections:
            return 0
        writable_rows = list(ACTION_ENTRY_ROWS)
        if len(report.action_items) > len(writable_rows):
            LOGGER.warning("PG Bible actions has %s contacts but only %s rows; extra contacts were not exported.", len(report.action_items), len(writable_rows))
        for row, item in zip(writable_rows, report.action_items):
            discovery_target = item.discovery_target_name_title or " ".join(part for part in [item.person_name, item.person_title] if part)
            nbm_booked = " ".join(part for part in [item.nbm_booked_date, item.nbm_booked_name_title] if part)
            self._write_nbm_target(ws, f"B{row}", item.related_nbm_target)
            self._write_coordinate(ws, f"C{row}", discovery_target)
            self._write_coordinate(ws, f"F{row}", self._yes_no(item.discovery_completed, default_no=True))
            self._write_coordinate(ws, f"G{row}", item.discovery_next_action or item.manager_notes or "No next action set")
            self._write_coordinate(ws, f"J{row}", nbm_booked)
            self._write_coordinate(ws, f"M{row}", self._yes_no(item.why_buy))
            self._write_coordinate(ws, f"N{row}", self._yes_no(item.exec_first, default_no=True))
            self._write_coordinate(ws, f"O{row}", self._yes_no(item.prep_with_manager))
            self._write_coordinate(ws, f"P{row}", self._yes_no(item.nbm_completed, default_no=True))
            self._write_coordinate(ws, f"Q{row}", item.nbm_next_action or item.discovery_next_action or "No next action set")
            self._write_coordinate(ws, f"T{row}", item.vo_value)
        return min(len(report.action_items), len(writable_rows))

    def _write_weekly_results(self, ws, report: OwnerReport) -> int:
        region = self.header_cache["PG RESULTS"]
        key = self.weekly_key
        if key is None:
            raise PGBibleExportError("WEEK_KEY_AMBIGUOUS", "The weekly key column has not been resolved.", [])
        slot_rows = self._weekly_slot_rows(ws, region)
        sorted_results = sorted(report.weekly_results, key=lambda item: self._normal_week_key(item.week_key, key.data_type))
        written = 0
        for index, item in enumerate(sorted_results):
            if index < len(slot_rows):
                row = slot_rows[index]
            else:
                row = self._insert_weekly_row(ws, region)
                slot_rows.append(row)
            self._write_value(ws.cell(row, key.column), self._coerce_week_key(item.week_key, key.data_type))
            mapping = {
                "vitos_sent": item.vitos_sent,
                "vitos_chased": item.vitos_chased,
                "discovery_booked": item.discovery_booked,
                "discovery_completed": item.discovery_completed,
                "nbms_booked": item.nbms_booked,
                "nbms_exec_firsts": item.nbms_exec_firsts,
                "nbms_completed": item.nbms_completed,
                "pipeline_generated_vo_count": item.pipeline_generated_vo_count,
                "pipeline_generated_value": item.pipeline_generated_value,
            }
            self._write_mapping(ws, region, row, mapping)
            written += 1
        return written

    def _clear_template_inputs(self, ws) -> None:
        for coordinate in ["F3", "F5", "L3"]:
            self._write_coordinate(ws, coordinate, None)
        # The account / SAAP booked date block contains template examples. There
        # is no approved PipeFlow data source for it yet, so exported workbooks
        # must clear O3:P6 instead of leaking placeholder account values.
        self._clear_range(ws, "O3", "P6")
        for row in PLAN_ENTRY_ROWS:
            for col in ("B", "D", "L", "M"):
                self._write_coordinate(ws, f"{col}{row}", None)
        for row in MONTH_PLAN_ROWS.values():
            for col in ("P", "S"):
                self._write_coordinate(ws, f"{col}{row}", None)
        for row in ACTION_ENTRY_ROWS:
            for col in ("B", "C", "F", "G", "J", "M", "N", "O", "P", "Q", "T"):
                self._write_coordinate(ws, f"{col}{row}", None)

    def _clear_table(self, ws, region: TableRegion) -> None:
        for row in range(region.start_row, region.end_row + 1):
            if self._row_contains_any(ws, row, SECTION_LABELS + ["Summary"]):
                break
            for col in set(region.columns.values()):
                cell = ws.cell(row, col)
                if not self._is_formula(cell):
                    self._write_value(cell, None)

    def _clear_weekly_rows(self, ws, region: TableRegion) -> None:
        for row in self._weekly_slot_rows(ws, region):
            for col in set(region.columns.values()):
                cell = ws.cell(row, col)
                if not self._is_formula(cell):
                    self._write_value(cell, None)

    def _weekly_data_rows(self, ws, region: TableRegion) -> list[int]:
        rows = []
        for row in range(region.start_row, region.end_row + 1):
            row_text = self._row_text(ws, row)
            if "Summary" in row_text:
                break
            if any(marker in row_text for marker in QUARTER_MARKERS):
                continue
            if ws.cell(row, region.columns["week_number"]).value not in (None, "") or ws.cell(row, region.columns["week_commencing"]).value not in (None, ""):
                rows.append(row)
        return rows

    def _weekly_slot_rows(self, ws, region: TableRegion) -> list[int]:
        rows = []
        for row in range(region.start_row, region.end_row + 1):
            row_text = self._row_text(ws, row)
            if "Summary" in row_text:
                break
            if any(marker in row_text for marker in QUARTER_MARKERS):
                continue
            rows.append(row)
        return rows

    def _insert_weekly_row(self, ws, region: TableRegion) -> int:
        weekly_rows = self._weekly_slot_rows(ws, region)
        if not weekly_rows:
            raise PGBibleExportError("WRITE_OUT_OF_BOUNDS", "No weekly data row exists to copy formatting from.", [region.name])
        insert_at = weekly_rows[-1] + 1
        ws.insert_rows(insert_at)
        self._copy_row_style(ws, weekly_rows[-1], insert_at)
        region.end_row += 1
        return insert_at

    def _ensure_capacity(self, ws, region: TableRegion, row_count: int) -> None:
        available = region.end_row - region.start_row + 1
        if row_count <= available:
            return
        insert_count = row_count - available
        insert_at = region.end_row + 1
        if self._row_contains_any(ws, insert_at, SECTION_LABELS):
            raise PGBibleExportError("WRITE_OUT_OF_BOUNDS", "The table does not have enough writable rows inside its boundary.", [region.name])
        for _ in range(insert_count):
            ws.insert_rows(insert_at)
            self._copy_row_style(ws, max(region.start_row, insert_at - 1), insert_at)
            region.end_row += 1

    def _write_mapping(self, ws, region: TableRegion, row: int, mapping: dict[str, Any]) -> None:
        if row > region.end_row:
            raise PGBibleExportError("WRITE_OUT_OF_BOUNDS", "A write would exceed the detected table boundary.", [region.name, str(row)])
        for key, value in mapping.items():
            if key in region.columns:
                self._write_value(ws.cell(row, region.columns[key]), value)

    def _write_value(self, cell, value: Any) -> None:
        if isinstance(cell, MergedCell):
            raise PGBibleExportError("WRITE_OUT_OF_BOUNDS", "A write targeted a non-anchor merged cell.", [cell.coordinate])
        if self._is_formula(cell):
            return
        if isinstance(value, Decimal):
            cell.value = float(value)
        else:
            cell.value = value

    def _yes_no(self, value: Any, default_no: bool = False) -> str:
        if value in (None, ""):
            return "No" if default_no else ""
        text = str(value).strip().casefold()
        if text in {"yes", "y", "true", "1"}:
            return "Yes"
        if text in {"no", "n", "false", "0"}:
            return "No"
        return str(value).strip()

    def _write_coordinate(self, ws, coordinate: str, value: Any) -> None:
        cell = ws[coordinate]
        if isinstance(cell, MergedCell):
            for merged in ws.merged_cells.ranges:
                if (cell.row, cell.column) in merged.cells:
                    cell = ws.cell(merged.min_row, merged.min_col)
                    break
        self._write_value(cell, value)

    def _write_nbm_target(self, ws, coordinate: str, value: Any) -> None:
        self._write_coordinate(ws, coordinate, value)
        anchor = self._anchor_cell(ws, coordinate)
        self._apply_nbm_target_style(anchor, value)

    def _anchor_cell(self, ws, coordinate: str):
        cell = ws[coordinate]
        if not isinstance(cell, MergedCell):
            return cell
        for merged in ws.merged_cells.ranges:
            if (cell.row, cell.column) in merged.cells:
                return ws.cell(merged.min_row, merged.min_col)
        return cell

    def _apply_nbm_target_style(self, cell, value: Any) -> None:
        # PG Bible target badges must mirror PG Progress: the target number is
        # the account PG Bible order and the colour is number % 12.
        if value in (None, ""):
            return
        try:
            colour_index = int(str(value).strip()) % 12
        except ValueError:
            colour_index = 0
        fill_colour, font_colour = NBM_COLOUR_PALETTE[colour_index]
        cell.fill = PatternFill(fill_type="solid", fgColor=fill_colour)
        cell.font = Font(
            name=cell.font.name,
            size=cell.font.sz,
            bold=True,
            italic=cell.font.italic,
            color=font_colour,
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _clear_range(self, ws, start_coordinate: str, end_coordinate: str) -> None:
        for row in ws[start_coordinate:end_coordinate]:
            for cell in row:
                if not isinstance(cell, MergedCell):
                    self._write_value(cell, None)

    def _coerce_week_key(self, value: Any, data_type: str):
        if data_type == "date" and isinstance(value, str):
            return datetime.fromisoformat(value).date()
        return value

    def _normal_week_key(self, value: Any, data_type: str) -> str:
        value = self._coerce_week_key(value, data_type)
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return str(value)

    def _self_check(self, ws, baseline: dict[str, Any]) -> None:
        current = self._structural_snapshot(ws)
        mismatches = []
        for key in ["merges", "column_widths", "row_heights", "header_strings"]:
            if current[key] != baseline[key]:
                mismatches.append(key)
        if mismatches:
            raise PGBibleExportError(
                "WRITE_OUT_OF_BOUNDS",
                "The exported sheet structure does not match the cloned template.",
                mismatches,
            )

    def _structural_snapshot(self, ws) -> dict[str, Any]:
        return {
            "merges": sorted(str(rng) for rng in ws.merged_cells.ranges),
            "column_widths": {col: ws.column_dimensions[get_column_letter(col)].width for col in range(1, ws.max_column + 1)},
            "row_heights": {row: ws.row_dimensions[row].height for row in range(1, ws.max_row + 1)},
            "header_strings": self._header_strings(ws),
        }

    def _header_strings(self, ws) -> dict[str, str]:
        protected_rows = set()
        for section in self.sections.values():
            protected_rows.add(section.row)
        for region in self.header_cache.values():
            protected_rows.update(region.header_rows)
        values = {}
        for row in protected_rows:
            for col in range(1, ws.max_column + 1):
                value = ws.cell(row, col).value
                if isinstance(value, str) and value.strip():
                    values[f"{row}:{col}"] = value
        return values

    def _copy_row_style(self, ws, source_row: int, target_row: int) -> None:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
        for col in range(1, ws.max_column + 1):
            source = ws.cell(source_row, col)
            target = ws.cell(target_row, col)
            if source.has_style:
                target._style = copy.copy(source._style)
            if source.number_format:
                target.number_format = source.number_format

    def _find_exact(self, ws, text: str):
        matches = []
        target = norm(text)
        for row in ws.iter_rows():
            for cell in row:
                if norm(cell.value) == target:
                    matches.append(cell)
        return matches

    def _merged_value(self, ws, row: int, col: int) -> Any:
        cell = ws.cell(row, col)
        if not isinstance(cell, MergedCell):
            return cell.value
        for merged in ws.merged_cells.ranges:
            if (row, col) in merged.cells:
                return ws.cell(merged.min_row, merged.min_col).value
        return None

    def _row_text(self, ws, row: int) -> str:
        return " ".join(str(ws.cell(row, col).value) for col in range(1, ws.max_column + 1) if ws.cell(row, col).value not in (None, ""))

    def _row_contains_any(self, ws, row: int, needles: list[str]) -> bool:
        text = self._row_text(ws, row)
        return any(needle in text for needle in needles)

    def _header_for_column(self, ws, region: TableRegion, col: int) -> str:
        for row in sorted(region.header_rows, reverse=True):
            value = self._merged_value(ws, row, col)
            if value not in (None, ""):
                return str(value)
        return get_column_letter(col)

    def _is_formula(self, cell) -> bool:
        return isinstance(cell.value, str) and cell.value.startswith("=")
